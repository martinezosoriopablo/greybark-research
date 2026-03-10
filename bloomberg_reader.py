# -*- coding: utf-8 -*-
"""
Greybark Research — Bloomberg Time Series Reader
=================================================

Reads the bloomberg_data.xlsx workbook (time series format v2.0)
and provides data to the AI Council agents and report generators.

Data sheet layout (as created by create_bloomberg_template.py):
  Row 1: Description
  Row 2: Campo ID
  Row 3: Bloomberg Ticker
  Row 4: BDH Field
  Row 5+: Col A = dates (YYYY-MM), Col B+ = values

Usage:
    from bloomberg_reader import BloombergData

    bbg = BloombergData()

    # Time series
    pmi = bbg.get_series("pmi_usa_mfg")   # pd.Series with DatetimeIndex
    latest = bbg.get_latest("pmi_usa_mfg") # float (most recent value)
    chg = bbg.get_change("pmi_usa_mfg", 1) # 1-month change

    # Full sheet as DataFrame
    df = bbg.get_sheet("PMI")

    # Formatted for AI agents
    text = bbg.format_for_macro_agent()

    # Backward compatible (v1 API)
    val = bbg.get("pmi_usa_mfg")
"""

import sys
import os
from pathlib import Path
from typing import Dict, Optional, Any, List
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None


# Data sheet layout constants (must match create_bloomberg_template.py)
DESC_ROW = 1
CAMPO_ROW = 2
TICKER_ROW = 3
FIELD_ROW = 4
DATA_START_ROW = 5

DATA_SHEETS = [
    # Template sheets (create_bloomberg_template.py)
    "PMI", "China", "CDS", "Credit_Spreads", "EM_Spreads",
    "Real_Yields", "CPI_Componentes", "EPFR_Flows", "Positioning",
    "Valuaciones", "Volatility", "Macro_Conditions", "Chile",
    "Factor_Returns", "Intl_Curves",
    # Actual file sheets (bloomberg_data.xlsx manual)
    "RF_Diaria", "Macro_Mensual", "RISK_DIARIA", "RV_Semanal",
    "Factors", "SOFR", "Indicators",
]


LIVE_BLOCK_SIZE = 135  # Must match create_bloomberg_template.BLOCK_SIZE


class BloombergData:
    """Reads Bloomberg time series data from the Excel template.

    Auto-detects format:
      - v2 (columnar): Row 1=desc, Row 2=campo_id across columns, Row 5+=data
      - v3-live (blocks): Series stacked vertically, BLOCK_SIZE rows apart,
        BDH formula fills dates in col A and values in col B
    """

    DEFAULT_PATHS = [
        Path(__file__).resolve().parent / "input" / "bloomberg_live.xlsx",
        Path(__file__).resolve().parent / "input" / "bloomberg_data.xlsm",
        Path(__file__).resolve().parent / "input" / "bloomberg_data.xlsx",
    ]

    def __init__(self, path: Optional[str] = None):
        if path:
            self.path = Path(path)
        else:
            self.path = None
            for p in self.DEFAULT_PATHS:
                if p.exists():
                    self.path = p
                    break

        self._series: Dict[str, 'pd.Series'] = {}
        self._meta: Dict[str, Dict[str, str]] = {}
        self._sheets: Dict[str, 'pd.DataFrame'] = {}
        self._loaded = False

        if self.path and self.path.exists():
            self._load()
        else:
            print(f"  [Bloomberg] Archivo no encontrado")

    def _detect_format(self, wb) -> str:
        """Detect template format: 'columnar' (v2) or 'blocks' (v3-live)."""
        if 'CONFIG' in wb.sheetnames:
            ws_cfg = wb['CONFIG']
            for r in range(2, 10):
                key = ws_cfg.cell(row=r, column=1).value
                val = ws_cfg.cell(row=r, column=2).value
                if key == 'formato' and val == 'bloques_verticales':
                    return 'blocks'
                if key == 'version_template' and val and 'live' in str(val):
                    return 'blocks'
        # Check first data sheet: if row 2 col 2 looks like a campo_id, it's columnar
        for sn in DATA_SHEETS:
            if sn in wb.sheetnames:
                ws = wb[sn]
                r2c2 = ws.cell(row=CAMPO_ROW, column=2).value
                if r2c2 and '_' in str(r2c2):
                    return 'columnar'
                r1c1 = ws.cell(row=1, column=1).value
                if r1c1 and '_' in str(r1c1):
                    return 'blocks'
                break
        return 'columnar'

    def _load(self):
        """Load all data sheets from the Excel workbook."""
        if not openpyxl:
            print("  [Bloomberg] openpyxl no instalado")
            return
        if not pd:
            print("  [Bloomberg] pandas no instalado")
            return

        wb = openpyxl.load_workbook(str(self.path), data_only=True)
        fmt = self._detect_format(wb)

        if fmt == 'blocks':
            self._load_blocks(wb)
        else:
            self._load_columnar(wb)

        wb.close()
        self._loaded = True

        # Supplement EM_Spreads from BCRP if Bloomberg sheet is empty
        em_sheet = self._sheets.get('EM_Spreads')
        em_empty = em_sheet is None or (em_sheet is not None and em_sheet.dropna(how='all').empty)
        has_embi = any(k.startswith('embi_') for k in self._series)
        if em_empty or not has_embi:
            self._supplement_from_bcrp()

        # Supplement with ECB data (always — unique series)
        self._supplement_from_ecb()

        n_series = len(self._series)
        n_points = sum(len(s) for s in self._series.values())
        if n_series > 0:
            print(f"  [Bloomberg] Cargado: {n_series} series, {n_points} data points")

    def _load_blocks(self, wb):
        """Load v3-live format: series in vertical blocks with BDH output."""
        for sheet_name in DATA_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            sheet_series = {}
            block = 0

            while True:
                base_row = block * LIVE_BLOCK_SIZE + 1

                # Read metadata row
                campo_id = ws.cell(row=base_row, column=1).value
                if not campo_id:
                    break
                campo_id = str(campo_id).strip()

                desc_raw = str(ws.cell(row=base_row, column=2).value or '')
                ticker = str(ws.cell(row=base_row, column=3).value or '')
                field = str(ws.cell(row=base_row, column=4).value or '')

                # Parse unit from description
                desc = desc_raw
                unit = ''
                if '(' in desc and desc.endswith(')'):
                    unit = desc[desc.rfind('(') + 1:-1]
                    desc = desc[:desc.rfind('(')].strip()

                self._meta[campo_id] = {
                    'description': desc,
                    'ticker': ticker,
                    'field': field,
                    'unit': unit,
                    'sheet': sheet_name,
                }

                # Read BDH output: col A = dates, col B = values (from base_row+1)
                dates = []
                values = []
                for r in range(base_row + 1, base_row + LIVE_BLOCK_SIZE):
                    date_val = ws.cell(row=r, column=1).value
                    data_val = ws.cell(row=r, column=2).value

                    if date_val is None:
                        break

                    dt = self._parse_date(date_val)
                    if dt is None:
                        continue

                    if data_val is not None and isinstance(data_val, (int, float)):
                        dates.append(dt)
                        values.append(float(data_val))

                if dates and values:
                    s = pd.Series(values, index=pd.DatetimeIndex(dates), name=campo_id)
                    s = s.sort_index()
                    self._series[campo_id] = s
                    sheet_series[campo_id] = s

                block += 1

            # Build sheet DataFrame
            if sheet_series:
                df = pd.DataFrame(sheet_series)
                df.index.name = 'date'
                df = df.sort_index()
                self._sheets[sheet_name] = df

    @staticmethod
    def _detect_row_layout(ws) -> dict:
        """Detect header row mapping for a sheet.

        Standard layout (create_bloomberg_template.py):
          Row 1: Description, Row 2: Campo ID, Row 3: Ticker, Row 4: Field

        Alternate layout (manual bloomberg_data.xlsx):
          Row 1: Section (or empty), Row 2: Campo ID / Short Name,
          Row 3: Description, Row 4: Bloomberg Ticker
          (No BDH field row — defaults to PX_LAST)
        """
        r4a = str(ws.cell(row=4, column=1).value or '').strip().lower()

        # If row 4 col A says "Bloomberg Ticker" → alternate layout
        if 'ticker' in r4a or 'bloomberg' in r4a:
            return {'desc': 3, 'campo': 2, 'ticker': 4, 'field': None}

        # Standard layout
        return {'desc': DESC_ROW, 'campo': CAMPO_ROW, 'ticker': TICKER_ROW, 'field': FIELD_ROW}

    def _load_columnar(self, wb):
        """Load v2 columnar format: headers in rows 1-4, data in columns."""
        for sheet_name in DATA_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            max_col = ws.max_column
            if max_col < 2:
                continue

            layout = self._detect_row_layout(ws)
            campo_row = layout['campo']
            desc_row = layout['desc']
            ticker_row = layout['ticker']
            field_row = layout['field']

            campo_ids = []
            descriptions = []
            tickers = []
            fields = []

            for col in range(2, max_col + 1):
                cid = ws.cell(row=campo_row, column=col).value
                if not cid:
                    break
                campo_ids.append(str(cid).strip())
                descriptions.append(str(ws.cell(row=desc_row, column=col).value or ''))
                tickers.append(str(ws.cell(row=ticker_row, column=col).value or ''))
                if field_row:
                    fields.append(str(ws.cell(row=field_row, column=col).value or ''))
                else:
                    fields.append('PX_LAST')

            if not campo_ids:
                continue

            dates = []
            data_rows = []
            empty_streak = 0
            for row_num in range(DATA_START_ROW, ws.max_row + 1):
                date_val = ws.cell(row=row_num, column=1).value
                if not date_val:
                    empty_streak += 1
                    if empty_streak > 5:
                        break
                    continue
                empty_streak = 0

                dt = self._parse_date(date_val)
                if dt is None:
                    continue

                dates.append(dt)
                row_values = []
                for col_idx in range(2, 2 + len(campo_ids)):
                    v = ws.cell(row=row_num, column=col_idx).value
                    if v is not None and isinstance(v, (int, float)):
                        row_values.append(float(v))
                    else:
                        row_values.append(None)
                data_rows.append(row_values)

            if not dates or not data_rows:
                continue

            df = pd.DataFrame(data_rows, index=pd.DatetimeIndex(dates), columns=campo_ids)
            df.index.name = 'date'
            df = df.sort_index()
            self._sheets[sheet_name] = df

            for i, cid in enumerate(campo_ids):
                series = df[cid].dropna()
                if len(series) > 0:
                    self._series[cid] = series

                desc = descriptions[i]
                unit = ''
                if '(' in desc and desc.endswith(')'):
                    unit = desc[desc.rfind('(') + 1:-1]
                    desc = desc[:desc.rfind('(')].strip()

                self._meta[cid] = {
                    'description': desc,
                    'ticker': tickers[i],
                    'field': fields[i],
                    'unit': unit,
                    'sheet': sheet_name,
                }

    @staticmethod
    def _parse_date(val) -> Optional[datetime]:
        """Parse a date value from Excel (datetime, string YYYY-MM, etc.)."""
        if isinstance(val, datetime):
            return val
        if hasattr(val, 'year'):
            return datetime(val.year, val.month, val.day)
        s = str(val).strip()
        # Try YYYY-MM
        if len(s) >= 7 and '-' in s:
            parts = s.split('-')
            try:
                return datetime(int(parts[0]), int(parts[1]), 1)
            except (ValueError, IndexError):
                pass
        # Try other formats
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y'):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    # ── BCRP EMBI supplement ──────────────────────────────────────────

    def _supplement_from_bcrp(self):
        """Fetch EMBI spreads from BCRP and inject into _series/_sheets."""
        try:
            from bcrp_embi_client import BCRPEmbiClient, EMBI_DESCRIPTIONS
        except ImportError:
            return

        client = BCRPEmbiClient()
        bcrp_series = client.fetch_embi_series()
        if not bcrp_series:
            return

        injected = 0
        sheet_series = {}
        for campo_id, s in bcrp_series.items():
            if campo_id not in self._series:
                self._series[campo_id] = s
                injected += 1
            sheet_series[campo_id] = self._series[campo_id]

            # Ensure metadata exists
            if campo_id not in self._meta:
                self._meta[campo_id] = {
                    'description': EMBI_DESCRIPTIONS.get(campo_id, campo_id),
                    'ticker': 'BCRP',
                    'field': 'EMBIG',
                    'unit': 'bps',
                    'sheet': 'EM_Spreads',
                }

        # Rebuild EM_Spreads sheet DataFrame
        if sheet_series:
            df = pd.DataFrame(sheet_series)
            df.index.name = 'date'
            df = df.sort_index()
            self._sheets['EM_Spreads'] = df

        if injected > 0:
            print(f"  [Bloomberg] BCRP EMBI: +{injected} series inyectadas")

    # ── ECB supplement ─────────────────────────────────────────────────

    def _supplement_from_ecb(self):
        """Fetch ECB macro series and inject as standalone values."""
        try:
            from ecb_client import ECBClient, ECB_DESCRIPTIONS
        except ImportError:
            return

        client = ECBClient()
        ecb_data = client.fetch_euro_macro()
        if not ecb_data:
            return

        injected = 0
        for campo_id, val in ecb_data.items():
            if val is None:
                continue
            if campo_id not in self._series:
                # Store as a single-point Series (latest value)
                s = pd.Series([val], index=pd.DatetimeIndex([datetime.now()]), name=campo_id)
                self._series[campo_id] = s
                injected += 1

            if campo_id not in self._meta:
                self._meta[campo_id] = {
                    'description': ECB_DESCRIPTIONS.get(campo_id, campo_id),
                    'ticker': 'ECB',
                    'field': 'SDMX',
                    'unit': '',
                    'sheet': 'Macro_Conditions',
                }

        if injected > 0:
            print(f"  [Bloomberg] ECB: +{injected} series inyectadas")

    # ── Primary API (Time Series) ───────────────────────────────────────

    def get_series(self, campo_id: str) -> Optional['pd.Series']:
        """Get full time series for a campo. Returns pd.Series with DatetimeIndex."""
        return self._series.get(campo_id)

    def get_latest(self, campo_id: str, default=None) -> Any:
        """Get the most recent value for a campo."""
        s = self._series.get(campo_id)
        if s is not None and len(s) > 0:
            return s.iloc[-1]
        return default

    def get_previous(self, campo_id: str, default=None) -> Any:
        """Get the second most recent value (previous month)."""
        s = self._series.get(campo_id)
        if s is not None and len(s) > 1:
            return s.iloc[-2]
        return default

    def get_change(self, campo_id: str, months: int = 1) -> Optional[float]:
        """Get the change over N months (latest - N months ago)."""
        s = self._series.get(campo_id)
        if s is None or len(s) <= months:
            return None
        return float(s.iloc[-1] - s.iloc[-1 - months])

    def get_pct_change(self, campo_id: str, months: int = 1) -> Optional[float]:
        """Get the percentage change over N months."""
        s = self._series.get(campo_id)
        if s is None or len(s) <= months:
            return None
        prev = s.iloc[-1 - months]
        if prev == 0:
            return None
        return float((s.iloc[-1] - prev) / abs(prev) * 100)

    def get_sheet(self, sheet_name: str) -> Optional['pd.DataFrame']:
        """Get the full DataFrame for a sheet."""
        return self._sheets.get(sheet_name)

    def get_meta(self, campo_id: str) -> Optional[Dict]:
        """Get metadata for a campo (description, ticker, field, unit, sheet)."""
        return self._meta.get(campo_id)

    def has(self, campo_id: str) -> bool:
        """Check if a campo has data."""
        return campo_id in self._series

    @property
    def available(self) -> bool:
        """Are Bloomberg data loaded?"""
        return self._loaded and len(self._series) > 0

    @property
    def campos(self) -> List[str]:
        """List all available campo IDs."""
        return list(self._series.keys())

    @property
    def sheets(self) -> List[str]:
        """List all loaded sheet names."""
        return list(self._sheets.keys())

    # ── Backward Compatible API (v1) ────────────────────────────────────

    def get(self, campo_id: str, default=None) -> Any:
        """Get latest value (v1 compatible)."""
        return self.get_latest(campo_id, default)

    def get_with_date(self, campo_id: str) -> Optional[Dict]:
        """Get latest value + date (v1 compatible)."""
        s = self._series.get(campo_id)
        if s is None or len(s) == 0:
            return None
        meta = self._meta.get(campo_id, {})
        return {
            'valor': s.iloc[-1],
            'fecha': s.index[-1],
            'unidad': meta.get('unit', ''),
            'seccion': meta.get('sheet', ''),
        }

    def section(self, seccion: str) -> Dict[str, float]:
        """Get all campos in a sheet/section as {campo_id: latest_value} (v1 compatible)."""
        result = {}
        for cid, meta in self._meta.items():
            if meta['sheet'] == seccion:
                val = self.get_latest(cid)
                if val is not None:
                    result[cid] = val
        return result

    def history(self, sheet_name: str) -> Optional[Any]:
        """Get DataFrame for a sheet (v1 compatible)."""
        return self.get_sheet(sheet_name)

    def has_section(self, seccion: str) -> bool:
        """Check if a section/sheet has data."""
        return seccion in self._sheets

    def get_cds_data(self) -> Dict[str, float]:
        """CDS data for RF report (v1 compatible)."""
        mapping = {
            'USA': 'cds_usa', 'Alemania': 'cds_alemania', 'UK': 'cds_uk',
            'Japon': 'cds_japon', 'Francia': 'cds_francia', 'Italia': 'cds_italia',
            'Espana': 'cds_espana', 'Brasil': 'cds_brasil', 'Mexico': 'cds_mexico',
            'Colombia': 'cds_colombia', 'Peru': 'cds_peru', 'Chile': 'cds_chile',
            'China': 'cds_china', 'Turquia': 'cds_turquia',
        }
        return {k: self.get_latest(v) for k, v in mapping.items() if self.has(v)}

    def get_sector_spreads(self) -> Dict[str, float]:
        """Sector spreads for RF report — IG and HY separated."""
        mapping = {
            # IG
            'Financiero IG': 'oas_ig_financiero', 'Industrial IG': 'oas_ig_industrial',
            'Utilities IG': 'oas_ig_utilities', 'Tecnologia IG': 'oas_ig_tecnologia',
            'Salud IG': 'oas_ig_salud', 'Energia IG': 'oas_ig_energia',
            'Total IG': 'oas_ig_total',
            # HY
            'Financiero HY': 'oas_hy_financiero', 'Industrial HY': 'oas_hy_industrial',
            'Utilities HY': 'oas_hy_utilities', 'Tecnologia HY': 'oas_hy_tecnologia',
            'Salud HY': 'oas_hy_salud', 'Energia HY': 'oas_hy_energia',
            'Total HY': 'oas_hy_total',
            # Legacy (template-generated)
            'Consumo': 'oas_consumo', 'Telecom': 'oas_telecom',
            'BBB': 'oas_bbb', 'BB': 'oas_bb',
            'IG Total': 'oas_ig_total_legacy', 'HY Total': 'oas_hy_total_legacy',
        }
        # Also try legacy campo_ids for backward compat
        legacy = {
            'Financiero': 'oas_financiero', 'Industrial': 'oas_industrial',
            'Utilities': 'oas_utilities', 'Tecnologia': 'oas_tecnologia',
            'Salud': 'oas_salud', 'Energia': 'oas_energia',
        }
        result = {k: self.get_latest(v) for k, v in mapping.items() if self.has(v)}
        for k, v in legacy.items():
            if k not in result and self.has(v):
                result[k] = self.get_latest(v)
        return result

    def get_pmi_latest(self) -> Dict[str, float]:
        """Latest PMIs for Macro report (v1 compatible)."""
        mapping = {
            'usa_mfg': 'pmi_usa_mfg', 'usa_svc': 'pmi_usa_svc',
            'euro_mfg': 'pmi_euro_mfg', 'euro_svc': 'pmi_euro_svc',
            'euro_comp': 'pmi_euro_comp', 'china_mfg': 'pmi_china_mfg',
            'china_caixin': 'pmi_china_caixin', 'japan': 'pmi_japan',
            'japan_comp': 'pmi_japan_comp',
            'global': 'pmi_global',
        }
        return {k: self.get_latest(v) for k, v in mapping.items() if self.has(v)}

    def get_epfr_flows(self) -> Dict[str, float]:
        """EPFR flows for RV report (v1 compatible)."""
        mapping = {
            'equity_usa': 'flujo_equity_usa', 'equity_europa': 'flujo_equity_europa',
            'equity_em': 'flujo_equity_em', 'equity_japan': 'flujo_equity_japan',
            'equity_latam': 'flujo_equity_latam',
            'bond_ig': 'flujo_bond_ig', 'bond_hy': 'flujo_bond_hy',
            'bond_em': 'flujo_bond_em',
        }
        return {k: self.get_latest(v) for k, v in mapping.items() if self.has(v)}

    def get_embi_spreads(self) -> Dict[str, float]:
        """EMBI spreads for AA report (v1 compatible)."""
        mapping = {
            'total': 'embi_total', 'Brasil': 'embi_brasil',
            'Mexico': 'embi_mexico', 'Colombia': 'embi_colombia',
            'Chile': 'embi_chile', 'Peru': 'embi_peru',
            'Argentina': 'embi_argentina', 'LatAm': 'embi_latam',
            'CEMBI': 'cembi_total',
        }
        return {k: self.get_latest(v) for k, v in mapping.items() if self.has(v)}

    def get_china_extended(self) -> Dict[str, Any]:
        """China extended data including property, credit, trade."""
        mapping = {
            'property_sales_yoy': 'china_property_sales_yoy',
            'tsf_yoy': 'china_tsf_yoy',
            'm2_yoy': 'china_m2_yoy',
            'new_loans': 'china_new_loans',
            'exports_yoy': 'china_exp_yoy',
            'imports_yoy': 'china_imp_yoy',
            'trade_balance': 'china_trade_bal',
            'ppi_yoy': 'china_ppi_yoy',
            'cpi_yoy': 'china_cpi_yoy',
            'caixin_mfg': 'china_caixin_mfg',
        }
        return {k: self.get_latest(v) for k, v in mapping.items() if self.has(v)}

    def get_valuations_extended(self) -> Dict[str, Dict[str, Any]]:
        """Extended valuations: PE fwd, PE 10Y avg, EV/EBITDA, Div Yield by region."""
        regions = {
            'spx': {'pe': 'pe_spx', 'pe_10y': 'pe_10y_spx', 'ev': 'ev_ebitda_spx', 'dy': 'dy_spx'},
            'stoxx600': {'pe': 'pe_stoxx600', 'pe_10y': 'pe_10y_stoxx600', 'ev': 'ev_ebitda_stoxx600', 'dy': 'dy_stoxx600'},
            'msci_em': {'pe': 'pe_msci_em', 'pe_10y': 'pe_10y_msci_em', 'ev': 'ev_ebitda_msci_em', 'dy': 'dy_msci_em'},
            'topix': {'pe': 'pe_topix', 'pe_10y': 'pe_10y_topix', 'ev': 'ev_ebitda_topix', 'dy': 'dy_topix'},
            'ipsa': {'pe': 'pe_ipsa', 'pe_10y': 'pe_10y_ipsa', 'ev': 'ev_ebitda_ipsa', 'dy': 'dy_ipsa'},
        }
        result = {}
        for region, campos in regions.items():
            data = {}
            for key, campo_id in campos.items():
                val = self.get_latest(campo_id)
                if val is not None:
                    data[key] = val
            if data:
                result[region] = data
        return result

    def get_factor_returns(self) -> Dict[str, float]:
        """Factor returns YTD for RV report.

        Bloomberg stores PX_LAST (index levels), not returns.
        We compute YTD % change from the first available value this year.
        """
        import pandas as pd

        mapping = {
            'quality': 'factor_quality',
            'momentum': 'factor_momentum',
            'value': 'factor_value',
            'growth': 'factor_growth',
            'size': 'factor_size',
        }
        result = {}
        current_year = pd.Timestamp.now().year
        for k, campo_id in mapping.items():
            s = self.get_series(campo_id)
            if s is None or len(s) == 0:
                continue
            latest = s.iloc[-1]
            # Find first value of current year
            ytd_start = s.loc[s.index >= f'{current_year}-01-01']
            if len(ytd_start) >= 2:
                start_val = ytd_start.iloc[0]
                if start_val and start_val != 0:
                    result[k] = round(((latest / start_val) - 1) * 100, 1)
                    continue
            # Fallback: if only 1 data point or can't compute, skip
        return result

    def get_sofr_curve(self) -> Dict[str, float]:
        """SOFR swap curve for RF report."""
        mapping = {
            'rate': 'sofr_rate', '1w': 'sofr_1w', '1m': 'sofr_1m',
            '3m': 'sofr_3m', '6m': 'sofr_6m', '1y': 'sofr_1y',
            '2y': 'sofr_2y', '3y': 'sofr_3y', '4y': 'sofr_4y',
            '5y': 'sofr_5y', '6y': 'sofr_6y', '7y': 'sofr_7y',
            '8y': 'sofr_8y', '9y': 'sofr_9y', '10y': 'sofr_10y',
            '15y': 'sofr_15y', '20y': 'sofr_20y', '25y': 'sofr_25y',
            '30y': 'sofr_30y',
        }
        return {k: self.get_latest(v) for k, v in mapping.items() if self.has(v)}

    def get_intl_curves(self) -> Dict[str, Dict[str, float]]:
        """International sovereign yield curves (Bund, Gilt, JGB)."""
        curves = {
            'bund': {'2y': 'bund_2y', '5y': 'bund_5y', '30y': 'bund_30y'},
            'gilt': {'2y': 'gilt_2y', '5y': 'gilt_5y', '30y': 'gilt_30y'},
            'jgb': {'2y': 'jgb_2y', '5y': 'jgb_5y', '30y': 'jgb_30y'},
        }
        result = {}
        for curve_name, tenors in curves.items():
            data = {}
            for tenor, campo_id in tenors.items():
                val = self.get_latest(campo_id)
                if val is not None:
                    data[tenor] = val
            if data:
                result[curve_name] = data
        return result

    # ── Agent Formatters ────────────────────────────────────────────────

    def _fmt_series_line(self, campo_id: str, label: str = None) -> str:
        """Format one series as: Label: value (prev: value, chg: +/-value)."""
        val = self.get_latest(campo_id)
        if val is None:
            return ''
        meta = self._meta.get(campo_id, {})
        label = label or meta.get('description', campo_id)
        unit = meta.get('unit', '')

        prev = self.get_previous(campo_id)
        chg = self.get_change(campo_id, 1)

        parts = [f"  {label}: {val:.1f}"]
        if prev is not None:
            parts.append(f"prev: {prev:.1f}")
        if chg is not None:
            sign = '+' if chg >= 0 else ''
            parts.append(f"chg: {sign}{chg:.1f}")

        line = parts[0]
        if len(parts) > 1:
            line += f" ({', '.join(parts[1:])})"
        if unit:
            line += f" {unit}"
        return line

    def _fmt_section(self, title: str, campos: List[str],
                     labels: Optional[Dict[str, str]] = None) -> str:
        """Format a section with title and series lines."""
        lines = [title]
        for cid in campos:
            if self.has(cid):
                label = labels.get(cid) if labels else None
                line = self._fmt_series_line(cid, label)
                if line:
                    lines.append(line)
        if len(lines) == 1:
            return ''
        return '\n'.join(lines)

    def format_for_macro_agent(self) -> str:
        """Format Bloomberg data for the Macro AI agent."""
        if not self.available:
            return ''

        sections = []

        # PMI
        pmi_campos = [f for f in self.campos if f.startswith('pmi_')]
        s = self._fmt_section('PMI Global:', pmi_campos)
        if s:
            sections.append(s)

        # China
        china_campos = [f for f in self.campos if f.startswith('china_')]
        s = self._fmt_section('China:', china_campos)
        if s:
            sections.append(s)

        # CPI
        cpi_campos = [f for f in self.campos if f.startswith('cpi_') or f.startswith('pce_')]
        s = self._fmt_section('CPI/PCE USA:', cpi_campos)
        if s:
            sections.append(s)

        # Macro conditions
        macro_campos = ['lei_usa', 'cfnai', 'nfib_optimism', 'umich_sentiment',
                        'ism_new_orders', 'adp_employment']
        s = self._fmt_section('Condiciones Macro:', macro_campos)
        if s:
            sections.append(s)

        # Real yields
        ry_campos = [f for f in self.campos if f.startswith('tips_') or f.startswith('breakeven_')]
        s = self._fmt_section('Real Yields & Breakevens:', ry_campos)
        if s:
            sections.append(s)

        # Chile
        chile_campos = [f for f in self.campos if f in
                        ('ipsa', 'usdclp', 'bcp_10y', 'chile_cpi_yoy', 'tpm_chile')]
        s = self._fmt_section('Chile:', chile_campos)
        if s:
            sections.append(s)

        # Europa / BCE
        eu_campos = ['ecb_dfr', 'hicp_euro_yoy', 'ea_10y_yield', 'eur_usd', 'm3_euro_stock']
        s = self._fmt_section('Europa / BCE:', eu_campos)
        if s:
            sections.append(s)

        if not sections:
            return ''
        return '=== BLOOMBERG DATA: MACRO ===\n' + '\n\n'.join(sections)

    def format_for_rv_agent(self) -> str:
        """Format Bloomberg data for the Renta Variable AI agent."""
        if not self.available:
            return ''

        sections = []

        # Valuaciones (PE Forward)
        pe_fwd_campos = [f for f in self.campos if f.startswith('pe_') and '_10y_' not in f]
        s = self._fmt_section('Valuaciones (PE Forward):', pe_fwd_campos)
        if s:
            sections.append(s)

        # Valuaciones (PE 10Y Average)
        pe_10y_campos = [f for f in self.campos if f.startswith('pe_10y_')]
        s = self._fmt_section('PE Promedio 10Y:', pe_10y_campos)
        if s:
            sections.append(s)

        # EV/EBITDA
        ev_campos = [f for f in self.campos if f.startswith('ev_ebitda_')]
        s = self._fmt_section('EV/EBITDA:', ev_campos)
        if s:
            sections.append(s)

        # Dividend Yields
        dy_campos = [f for f in self.campos if f.startswith('dy_')]
        s = self._fmt_section('Dividend Yield:', dy_campos)
        if s:
            sections.append(s)

        # Factor Returns
        factor_campos = [f for f in self.campos if f.startswith('factor_')]
        s = self._fmt_section('Factor Returns YTD:', factor_campos)
        if s:
            sections.append(s)

        # EPFR Flows
        flow_campos = [f for f in self.campos if f.startswith('flujo_')]
        s = self._fmt_section('Flujos EPFR:', flow_campos)
        if s:
            sections.append(s)

        # Positioning
        pos_campos = [f for f in self.campos if f in
                      ('aaii_bullish', 'aaii_bearish', 'put_call_ratio',
                       'naaim_exposure', 'inv_intel_bulls', 'inv_intel_bears')]
        s = self._fmt_section('Posicionamiento:', pos_campos)
        if s:
            sections.append(s)

        # Volatility
        vol_campos = [f for f in self.campos if f in ('vix', 'vxeem', 'skew', 'vix_3m')]
        s = self._fmt_section('Volatilidad:', vol_campos)
        if s:
            sections.append(s)

        if not sections:
            return ''
        return '=== BLOOMBERG DATA: RENTA VARIABLE ===\n' + '\n\n'.join(sections)

    def format_for_rf_agent(self) -> str:
        """Format Bloomberg data for the Renta Fija AI agent."""
        if not self.available:
            return ''

        sections = []

        # SOFR Swap Curve
        sofr_campos = ['sofr_rate', 'sofr_1y', 'sofr_2y', 'sofr_3y',
                        'sofr_5y', 'sofr_7y', 'sofr_10y', 'sofr_20y', 'sofr_30y']
        s = self._fmt_section('SOFR Swap Curve:', sofr_campos)
        if s:
            sections.append(s)

        # International Sovereign Curves (Bund, Gilt, JGB)
        bund_campos = ['bund_2y', 'bund_5y', 'bund_30y']
        s = self._fmt_section('German Bund:', bund_campos)
        if s:
            sections.append(s)

        gilt_campos = ['gilt_2y', 'gilt_5y', 'gilt_30y']
        s = self._fmt_section('UK Gilt:', gilt_campos)
        if s:
            sections.append(s)

        jgb_campos = ['jgb_2y', 'jgb_5y', 'jgb_30y']
        s = self._fmt_section('JGB:', jgb_campos)
        if s:
            sections.append(s)

        # CDS
        cds_campos = [f for f in self.campos if f.startswith('cds_')]
        s = self._fmt_section('CDS Soberanos 5Y:', cds_campos)
        if s:
            sections.append(s)

        # Credit Spreads IG
        oas_ig = [f for f in self.campos if f.startswith('oas_ig_')]
        s = self._fmt_section('OAS Investment Grade:', oas_ig)
        if s:
            sections.append(s)

        # Credit Spreads HY
        oas_hy = [f for f in self.campos if f.startswith('oas_hy_')]
        s = self._fmt_section('OAS High Yield:', oas_hy)
        if s:
            sections.append(s)

        # Legacy OAS (template-generated, no IG/HY prefix)
        oas_legacy = [f for f in self.campos if f.startswith('oas_')
                      and not f.startswith('oas_ig_') and not f.startswith('oas_hy_')]
        s = self._fmt_section('Spreads Credito:', oas_legacy)
        if s:
            sections.append(s)

        # EM Spreads
        em_campos = [f for f in self.campos if f.startswith('embi_') or f.startswith('cembi_')]
        s = self._fmt_section('EM Spreads:', em_campos)
        if s:
            sections.append(s)

        # Real Yields
        ry_campos = [f for f in self.campos if f.startswith('tips_') or
                     f.startswith('real_yield_') or f.startswith('breakeven_')]
        s = self._fmt_section('Real Yields & Breakevens:', ry_campos)
        if s:
            sections.append(s)

        if not sections:
            return ''
        return '=== BLOOMBERG DATA: RENTA FIJA ===\n' + '\n\n'.join(sections)

    def format_for_risk_agent(self) -> str:
        """Format Bloomberg data for the Riesgo AI agent."""
        if not self.available:
            return ''

        sections = []

        # Volatility
        vol_campos = [f for f in self.campos if f in
                      ('vix', 'move', 'vxeem', 'skew', 'vix_3m', 'v2x', 'gvz')]
        s = self._fmt_section('Volatilidad:', vol_campos)
        if s:
            sections.append(s)

        # CDS (systemic risk)
        cds_key = ['cds_usa', 'cds_alemania', 'cds_italia', 'cds_china', 'cds_turquia']
        s = self._fmt_section('CDS Soberanos (riesgo sistemico):', cds_key)
        if s:
            sections.append(s)

        # Credit (risk)
        credit_key = ['oas_ig_total', 'oas_hy_total', 'oas_bb']
        s = self._fmt_section('Spreads Credito:', credit_key)
        if s:
            sections.append(s)

        # Positioning (contrarian)
        pos_campos = ['aaii_bullish', 'aaii_bearish', 'put_call_ratio']
        s = self._fmt_section('Posicionamiento (senales contrarian):', pos_campos)
        if s:
            sections.append(s)

        if not sections:
            return ''
        return '=== BLOOMBERG DATA: RIESGO ===\n' + '\n\n'.join(sections)

    def format_for_geo_agent(self) -> str:
        """Format Bloomberg data for the Geopolitica AI agent."""
        if not self.available:
            return ''

        sections = []

        # China trade/credit
        china_campos = ['china_exp_yoy', 'china_imp_yoy', 'china_trade_bal',
                        'china_tsf_yoy', 'china_iron_ore', 'china_copper']
        s = self._fmt_section('China (comercio & commodities):', china_campos)
        if s:
            sections.append(s)

        # EM CDS (geopolitical stress)
        em_cds = ['cds_brasil', 'cds_mexico', 'cds_chile', 'cds_china', 'cds_turquia']
        s = self._fmt_section('CDS EM (stress geopolitico):', em_cds)
        if s:
            sections.append(s)

        # EMBI LatAm
        embi_geo = ['embi_latam', 'embi_brasil', 'embi_mexico', 'embi_colombia',
                     'embi_chile', 'embi_peru', 'embi_argentina']
        s = self._fmt_section('EMBI LatAm (riesgo soberano):', embi_geo)
        if s:
            sections.append(s)

        # Chile
        chile_campos = ['ipsa', 'usdclp', 'bcp_10y']
        s = self._fmt_section('Chile:', chile_campos)
        if s:
            sections.append(s)

        if not sections:
            return ''
        return '=== BLOOMBERG DATA: GEOPOLITICA ===\n' + '\n\n'.join(sections)

    # ── Summary ─────────────────────────────────────────────────────────

    def summary(self):
        """Print summary of loaded data."""
        if not self._series:
            print("  [Bloomberg] Sin datos cargados")
            return

        print(f"\n  Bloomberg Data Summary ({len(self._series)} series):")
        for sheet_name in DATA_SHEETS:
            if sheet_name in self._sheets:
                df = self._sheets[sheet_name]
                n_points = df.notna().sum().sum()
                print(f"    {sheet_name}: {len(df.columns)} series, {int(n_points)} data points")

        # Date range
        all_dates = set()
        for s in self._series.values():
            all_dates.update(s.index)
        if all_dates:
            print(f"  Rango: {min(all_dates).strftime('%Y-%m')} a {max(all_dates).strftime('%Y-%m')}")


# ═══════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    bbg = BloombergData()
    bbg.summary()

    if bbg.available:
        print("\n--- Backward Compatible API ---")
        print("CDS:", bbg.get_cds_data())
        print("PMI:", bbg.get_pmi_latest())
        print("Sectores:", bbg.get_sector_spreads())
        print("SOFR:", bbg.get_sofr_curve())
        print("EPFR:", bbg.get_epfr_flows())
        print("EMBI:", bbg.get_embi_spreads())

        print("\n--- Agent Formatters ---")
        for name, fmt_fn in [
            ("Macro", bbg.format_for_macro_agent),
            ("RV", bbg.format_for_rv_agent),
            ("RF", bbg.format_for_rf_agent),
            ("Riesgo", bbg.format_for_risk_agent),
            ("Geo", bbg.format_for_geo_agent),
        ]:
            text = fmt_fn()
            if text:
                print(f"\n{text[:500]}...")
            else:
                print(f"\n[{name}] Sin datos")
    else:
        print("\nLlena el Excel en input/bloomberg_data.xlsx y vuelve a ejecutar.")
