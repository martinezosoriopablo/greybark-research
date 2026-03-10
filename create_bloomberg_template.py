# -*- coding: utf-8 -*-
"""
Greybark Research — Bloomberg Template Generator (Time Series)
==============================================================

Creates the Bloomberg Excel workbook with time series format.
Each sheet has 24 months of monthly data as pure values.

Architecture:
  - CONFIG sheet with metadata
  - INSTRUCCIONES sheet
  - 15 thematic data sheets (pure values, safe on non-Bloomberg PCs)
  - Catalogo reference sheet
  - VBA module (.bas) for one-click Bloomberg refresh

Data sheet layout:
  Row 1: Description (human-readable header)
  Row 2: Campo ID (for Python code)
  Row 3: Bloomberg Ticker (for VBA macro)
  Row 4: BDH Field (PX_LAST, BEST_PE_RATIO, etc.)
  Row 5+: Col A = dates (YYYY-MM), Col B+ = values

Usage:
    python create_bloomberg_template.py
"""

import sys
import os
from pathlib import Path
from datetime import date

sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Estilos ──────────────────────────────────────────────────────────────
ORANGE = "DD6B20"
BLACK = "1A1A1A"
WHITE = "FFFFFF"
LIGHT_GRAY = "F7F7F7"
MED_GRAY = "E0E0E0"
DARK_GRAY = "888888"

header_font = Font(name="Segoe UI", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=BLACK, end_color=BLACK, fill_type="solid")
title_font = Font(name="Segoe UI", bold=True, color=ORANGE, size=14)
subtitle_font = Font(name="Segoe UI", bold=True, color=BLACK, size=11)
normal_font = Font(name="Segoe UI", size=10)
small_gray_font = Font(name="Segoe UI", size=8, color=DARK_GRAY)
note_font = Font(name="Segoe UI", italic=True, color="666666", size=9)
date_font = Font(name="Segoe UI", size=10, color=BLACK)
thin_border = Border(
    bottom=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
)
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

MONTHS_HISTORY = 120   # 10 years of monthly data
MONTHS_FORWARD = 3     # Buffer: rows for future months so template lasts longer
DATA_START_ROW = 5


# ═════════════════════════════════════════════════════════════════════════
# CATALOGO: 148 series across 15 sheets
# Format: (sheet, campo_id, ticker, field, description, unit)
# ═════════════════════════════════════════════════════════════════════════

CATALOG = [
    # ── PMI (11 series) ─────────────────────────────────────────────────
    ("PMI", "pmi_usa_mfg", "NAPMPMI Index", "PX_LAST", "ISM Manufacturing USA", "indice"),
    ("PMI", "pmi_usa_svc", "NAPMNMI Index", "PX_LAST", "ISM Services USA", "indice"),
    ("PMI", "pmi_euro_mfg", "MPMIezma Index", "PX_LAST", "PMI Mfg Eurozona", "indice"),
    ("PMI", "pmi_euro_svc", "MPMiezsa Index", "PX_LAST", "PMI Svc Eurozona", "indice"),
    ("PMI", "pmi_euro_comp", "MPMIEZCa Index", "PX_LAST", "PMI Composite Eurozona", "indice"),
    ("PMI", "pmi_china_mfg", "CPMINDX Index", "PX_LAST", "PMI Mfg China (NBS)", "indice"),
    ("PMI", "pmi_china_caixin", "CPMICOMP Index", "PX_LAST", "Caixin Composite China", "indice"),
    ("PMI", "pmi_japan", "MPMIJPMA Index", "PX_LAST", "PMI Mfg Japon", "indice"),
    ("PMI", "pmi_japan_comp", "MPMIJPCA Index", "PX_LAST", "PMI Composite Japon", "indice"),
    ("PMI", "pmi_uk", "PMITMUK Index", "PX_LAST", "PMI Mfg UK", "indice"),
    ("PMI", "pmi_india", "NPIMINM Index", "PX_LAST", "PMI Mfg India", "indice"),

    # ── China (12 series) ───────────────────────────────────────────────
    ("China", "china_exp_yoy", "CNFREXPY Index", "PX_LAST", "Exportaciones YoY", "%"),
    ("China", "china_imp_yoy", "CNFRIMPY Index", "PX_LAST", "Importaciones YoY", "%"),
    ("China", "china_trade_bal", "CNTSTCN Index", "PX_LAST", "Balanza Comercial", "USD bn"),
    ("China", "china_tsf_yoy", "CNTSFTOT Index", "PX_LAST", "Total Social Financing YoY", "%"),
    ("China", "china_m2_yoy", "CNMS2YOY Index", "PX_LAST", "M2 Money Supply YoY", "%"),
    ("China", "china_ppi_yoy", "CHEFTYOY Index", "PX_LAST", "PPI YoY", "%"),
    ("China", "china_cpi_yoy", "CNCPIYOY Index", "PX_LAST", "CPI YoY", "%"),
    ("China", "china_new_loans", "CNNELOYN Index", "PX_LAST", "New Yuan Loans", "CNY bn"),
    ("China", "china_property_sales_yoy", "CNPRSALY Index", "PX_LAST", "Property Sales YoY", "%"),
    ("China", "china_iron_ore", "SCO1 Comdty", "PX_LAST", "Iron Ore 62% Fe", "USD/t"),
    ("China", "china_copper", "HG1 Comdty", "PX_LAST", "Copper Future", "USD/lb"),
    ("China", "china_caixin_mfg", "CHPMINDX Index", "PX_LAST", "Caixin Mfg PMI", "indice"),

    # ── CDS Soberanos 5Y (14 series) ────────────────────────────────────
    ("CDS", "cds_usa", "CT786916 Corp", "PX_LAST", "USA CDS 5Y", "bps"),
    ("CDS", "cds_alemania", "cdbr1u5 corp", "PX_LAST", "Alemania CDS 5Y", "bps"),
    ("CDS", "cds_uk", "cUKT1u5 corp", "PX_LAST", "UK CDS 5Y", "bps"),
    ("CDS", "cds_japon", "CJGB1U5 CBGN Curncy", "PX_LAST", "Japon CDS 5Y", "bps"),
    ("CDS", "cds_francia", "CFRTR1U5 CORP", "PX_LAST", "Francia CDS 5Y", "bps"),
    ("CDS", "cds_italia", "CITLY1U5 CORP", "PX_LAST", "Italia CDS 5Y", "bps"),
    ("CDS", "cds_espana", "CSPA1U5 CORP", "PX_LAST", "Espana CDS 5Y", "bps"),
    ("CDS", "cds_brasil", "CBRZ1U5 CORP", "PX_LAST", "Brasil CDS 5Y", "bps"),
    ("CDS", "cds_mexico", "CMEX1U5 CORP", "PX_LAST", "Mexico CDS 5Y", "bps"),
    ("CDS", "cds_colombia", "CCOL1U5 CORP", "PX_LAST", "Colombia CDS 5Y", "bps"),
    ("CDS", "cds_peru", "CPERU1U5 CORP", "PX_LAST", "Peru CDS 5Y", "bps"),
    ("CDS", "cds_chile", "CCHIL1U5 CORP", "PX_LAST", "Chile CDS 5Y", "bps"),
    ("CDS", "cds_china", "CCHIN1U5 CORP", "PX_LAST", "China CDS 5Y", "bps"),
    ("CDS", "cds_turquia", "CTURK1U5 CORP", "PX_LAST", "Turquia CDS 5Y", "bps"),

    # ── Credit Spreads (22 series) ──────────────────────────────────────
    # IG Sector
    ("Credit_Spreads", "oas_ig_total", "USOAIGTO Index", "PX_LAST", "OAS IG Total (US)", "bps"),
    ("Credit_Spreads", "oas_ig_financiero", "USOAIGFI Index", "PX_LAST", "OAS IG Financiero", "bps"),
    ("Credit_Spreads", "oas_ig_industrial", "USOAIGIN Index", "PX_LAST", "OAS IG Industrial", "bps"),
    ("Credit_Spreads", "oas_ig_utilities", "USOAIGUT Index", "PX_LAST", "OAS IG Utilities", "bps"),
    ("Credit_Spreads", "oas_ig_tecnologia", "USOAIGTC Index", "PX_LAST", "OAS IG Tecnologia", "bps"),
    ("Credit_Spreads", "oas_ig_salud", "USOAIGHC Index", "PX_LAST", "OAS IG Salud", "bps"),
    ("Credit_Spreads", "oas_ig_energia", "USOAIGEN Index", "PX_LAST", "OAS IG Energia", "bps"),
    # HY Sector
    ("Credit_Spreads", "oas_hy_total", "USOHHYTO Index", "PX_LAST", "OAS HY Total (US)", "bps"),
    ("Credit_Spreads", "oas_hy_financiero", "USOHHYFI Index", "PX_LAST", "OAS HY Financiero", "bps"),
    ("Credit_Spreads", "oas_hy_industrial", "USOHHYIN Index", "PX_LAST", "OAS HY Industrial", "bps"),
    ("Credit_Spreads", "oas_hy_utilities", "USOHHYUT Index", "PX_LAST", "OAS HY Utilities", "bps"),
    ("Credit_Spreads", "oas_hy_tecnologia", "USOHHYTC Index", "PX_LAST", "OAS HY Tecnologia", "bps"),
    ("Credit_Spreads", "oas_hy_salud", "USOHHYHC Index", "PX_LAST", "OAS HY Salud", "bps"),
    ("Credit_Spreads", "oas_hy_energia", "USOHHYEN Index", "PX_LAST", "OAS HY Energia", "bps"),
    # Aggregate / Ratings
    ("Credit_Spreads", "oas_bbb", "LUACBBB Index", "PX_LAST", "OAS BBB", "bps"),
    ("Credit_Spreads", "oas_bb", "LF98BB Index", "PX_LAST", "OAS BB", "bps"),
    ("Credit_Spreads", "oas_consumo", "LUACCS Index", "PX_LAST", "OAS Consumo", "bps"),
    ("Credit_Spreads", "oas_telecom", "LUACTM Index", "PX_LAST", "OAS Telecom", "bps"),
    ("Credit_Spreads", "oas_ig_eur", "ER00OAS Index", "PX_LAST", "OAS IG Europa", "bps"),

    # ── EM Spreads (8 series) ───────────────────────────────────────────
    ("EM_Spreads", "embi_total", "JPEMCOMP Index", "PX_LAST", "EMBI+ Composite", "bps"),
    ("EM_Spreads", "embi_brasil", "JPEBBR Index", "PX_LAST", "EMBI Brasil", "bps"),
    ("EM_Spreads", "embi_mexico", "JPEBMX Index", "PX_LAST", "EMBI Mexico", "bps"),
    ("EM_Spreads", "embi_colombia", "JPEBCO Index", "PX_LAST", "EMBI Colombia", "bps"),
    ("EM_Spreads", "embi_chile", "JPEBCL Index", "PX_LAST", "EMBI Chile", "bps"),
    ("EM_Spreads", "cembi_total", "JCMBCOMP Index", "PX_LAST", "CEMBI Composite", "bps"),
    ("EM_Spreads", "embi_asia", "JPEIASIA Index", "PX_LAST", "EMBI Asia", "bps"),
    ("EM_Spreads", "embi_emea", "JPEEMEA Index", "PX_LAST", "EMBI EMEA", "bps"),

    # ── Real Yields & Breakevens (8 series) ─────────────────────────────
    ("Real_Yields", "tips_5y", "H15T5Y Index", "PX_LAST", "US TIPS 5Y Real Yield", "%"),
    ("Real_Yields", "tips_10y", "H15T10Y Index", "PX_LAST", "US TIPS 10Y Real Yield", "%"),
    ("Real_Yields", "real_yield_de_10y", "GTDEMII10YR Govt", "PX_LAST", "Bund Real Yield 10Y", "%"),
    ("Real_Yields", "real_yield_uk_10y", "GTGBPII10Y GOVT", "PX_LAST", "Gilt Real Yield 10Y", "%"),
    ("Real_Yields", "breakeven_us_10y", "USGGBE10 Index", "PX_LAST", "Breakeven US 10Y", "%"),
    ("Real_Yields", "breakeven_us_5y", "USGGBE05 Index", "PX_LAST", "Breakeven US 5Y", "%"),
    ("Real_Yields", "breakeven_de_10y", "DEGGBE10 Index", "PX_LAST", "Breakeven Alemania 10Y", "%"),
    ("Real_Yields", "breakeven_uk_10y", "UKGGBE10 Index", "PX_LAST", "Breakeven UK 10Y", "%"),

    # ── CPI Componentes USA (8 series) ──────────────────────────────────
    ("CPI_Componentes", "cpi_headline_yoy", "CPI YOY Index", "PX_LAST", "CPI Headline YoY", "%"),
    ("CPI_Componentes", "cpi_core_yoy", "CPI XYOY Index", "PX_LAST", "CPI Core YoY", "%"),
    ("CPI_Componentes", "pce_headline_yoy", "PCE DEFY Index", "PX_LAST", "PCE Headline YoY", "%"),
    ("CPI_Componentes", "pce_core_yoy", "PCE CYOY Index", "PX_LAST", "PCE Core YoY", "%"),
    ("CPI_Componentes", "cpi_shelter_yoy", "CPSHSHLT Index", "PX_LAST", "CPI Shelter YoY", "%"),
    ("CPI_Componentes", "cpi_energy_yoy", "CPI ENRG Index", "PX_LAST", "CPI Energy YoY", "%"),
    ("CPI_Componentes", "cpi_svc_ex_housing", "CPUPAXFE Index", "PX_LAST", "CPI Svc ex Housing YoY", "%"),
    ("CPI_Componentes", "cpi_goods_yoy", "CPUPCXFE Index", "PX_LAST", "CPI Goods YoY", "%"),

    # ── EPFR Flows (8 series) ───────────────────────────────────────────
    ("EPFR_Flows", "flujo_equity_usa", "EPFRUSEA Index", "PX_LAST", "EPFR Equity USA", "USD mn"),
    ("EPFR_Flows", "flujo_equity_europa", "EPFREQEU Index", "PX_LAST", "EPFR Equity Europa", "USD mn"),
    ("EPFR_Flows", "flujo_equity_em", "EPFREQEM Index", "PX_LAST", "EPFR Equity EM", "USD mn"),
    ("EPFR_Flows", "flujo_equity_japan", "EPFREQJP Index", "PX_LAST", "EPFR Equity Japon", "USD mn"),
    ("EPFR_Flows", "flujo_equity_latam", "EPFREQLA Index", "PX_LAST", "EPFR Equity LatAm", "USD mn"),
    ("EPFR_Flows", "flujo_bond_ig", "EPFRBDIG Index", "PX_LAST", "EPFR Bond IG", "USD mn"),
    ("EPFR_Flows", "flujo_bond_hy", "EPFRBDHY Index", "PX_LAST", "EPFR Bond HY", "USD mn"),
    ("EPFR_Flows", "flujo_bond_em", "EPFRBDEM Index", "PX_LAST", "EPFR Bond EM", "USD mn"),

    # ── Positioning (6 series) ──────────────────────────────────────────
    ("Positioning", "aaii_bullish", "AAIIBULL Index", "PX_LAST", "AAII % Bullish", "%"),
    ("Positioning", "aaii_bearish", "AAIIBEAR Index", "PX_LAST", "AAII % Bearish", "%"),
    ("Positioning", "put_call_ratio", "PCUSEQTR Index", "PX_LAST", "Put/Call Ratio Equity", "ratio"),
    ("Positioning", "naaim_exposure", "NAAIM Index", "PX_LAST", "NAAIM Exposure", "%"),
    ("Positioning", "inv_intel_bulls", "INELBULL Index", "PX_LAST", "Investors Intel Bulls", "%"),
    ("Positioning", "inv_intel_bears", "INELBEAR Index", "PX_LAST", "Investors Intel Bears", "%"),

    # ── Valuaciones (30 series) ─────────────────────────────────────────
    # PE Forward by region/sector
    ("Valuaciones", "pe_spx", "SPX Index", "BEST_PE_RATIO", "PE Fwd S&P 500", "x"),
    ("Valuaciones", "pe_ndx", "NDX Index", "BEST_PE_RATIO", "PE Fwd Nasdaq 100", "x"),
    ("Valuaciones", "pe_stoxx600", "SXXP Index", "BEST_PE_RATIO", "PE Fwd STOXX 600", "x"),
    ("Valuaciones", "pe_topix", "TPX Index", "BEST_PE_RATIO", "PE Fwd TOPIX", "x"),
    ("Valuaciones", "pe_msci_em", "MXEF Index", "BEST_PE_RATIO", "PE Fwd MSCI EM", "x"),
    ("Valuaciones", "pe_ipsa", "IPSA Index", "BEST_PE_RATIO", "PE Fwd IPSA", "x"),
    ("Valuaciones", "pe_tech", "S5INFT Index", "BEST_PE_RATIO", "PE Fwd Tech (S&P)", "x"),
    ("Valuaciones", "pe_healthcare", "S5HLTH Index", "BEST_PE_RATIO", "PE Fwd Healthcare", "x"),
    ("Valuaciones", "pe_financials", "S5FINL Index", "BEST_PE_RATIO", "PE Fwd Financials", "x"),
    ("Valuaciones", "pe_energy", "S5ENRS Index", "BEST_PE_RATIO", "PE Fwd Energy", "x"),
    ("Valuaciones", "pe_industrials", "S5INDU Index", "BEST_PE_RATIO", "PE Fwd Industrials", "x"),
    ("Valuaciones", "pe_consumer_disc", "S5COND Index", "BEST_PE_RATIO", "PE Fwd Consumer Disc", "x"),
    ("Valuaciones", "pe_consumer_stap", "S5CONS Index", "BEST_PE_RATIO", "PE Fwd Consumer Stap", "x"),
    ("Valuaciones", "pe_comm_svcs", "S5TELS Index", "BEST_PE_RATIO", "PE Fwd Comm Services", "x"),
    ("Valuaciones", "pe_materials", "S5MATR Index", "BEST_PE_RATIO", "PE Fwd Materials", "x"),
    # P/E 10Y Average by region (BEst LT PE Ratio = consensus 10Y trailing avg)
    ("Valuaciones", "pe_10y_spx", "SPX Index", "BEST_PE_RATIO_10YR_AVG", "PE 10Y Avg S&P 500", "x"),
    ("Valuaciones", "pe_10y_stoxx600", "SXXP Index", "BEST_PE_RATIO_10YR_AVG", "PE 10Y Avg STOXX 600", "x"),
    ("Valuaciones", "pe_10y_msci_em", "MXEF Index", "BEST_PE_RATIO_10YR_AVG", "PE 10Y Avg MSCI EM", "x"),
    ("Valuaciones", "pe_10y_topix", "TPX Index", "BEST_PE_RATIO_10YR_AVG", "PE 10Y Avg TOPIX", "x"),
    ("Valuaciones", "pe_10y_ipsa", "IPSA Index", "BEST_PE_RATIO_10YR_AVG", "PE 10Y Avg IPSA", "x"),
    # EV/EBITDA by region
    ("Valuaciones", "ev_ebitda_spx", "SPX Index", "BEST_EV_TO_BEST_EBITDA", "EV/EBITDA S&P 500", "x"),
    ("Valuaciones", "ev_ebitda_stoxx600", "SXXP Index", "BEST_EV_TO_BEST_EBITDA", "EV/EBITDA STOXX 600", "x"),
    ("Valuaciones", "ev_ebitda_msci_em", "MXEF Index", "BEST_EV_TO_BEST_EBITDA", "EV/EBITDA MSCI EM", "x"),
    ("Valuaciones", "ev_ebitda_topix", "TPX Index", "BEST_EV_TO_BEST_EBITDA", "EV/EBITDA TOPIX", "x"),
    ("Valuaciones", "ev_ebitda_ipsa", "IPSA Index", "BEST_EV_TO_BEST_EBITDA", "EV/EBITDA IPSA", "x"),
    # Dividend Yield (STOXX 600 was missing)
    ("Valuaciones", "dy_spx", "SPX Index", "BEST_DIV_YLD", "Div Yield S&P 500", "%"),
    ("Valuaciones", "dy_stoxx600", "SXXP Index", "BEST_DIV_YLD", "Div Yield STOXX 600", "%"),
    ("Valuaciones", "dy_topix", "TPX Index", "BEST_DIV_YLD", "Div Yield TOPIX", "%"),
    ("Valuaciones", "dy_msci_em", "MXEF Index", "BEST_DIV_YLD", "Div Yield MSCI EM", "%"),
    ("Valuaciones", "dy_ipsa", "IPSA Index", "BEST_DIV_YLD", "Div Yield IPSA", "%"),

    # ── Volatility (7 series) ───────────────────────────────────────────
    ("Volatility", "vix", "VIX Index", "PX_LAST", "VIX (S&P 500 Vol)", "indice"),
    ("Volatility", "move", "MOVE Index", "PX_LAST", "MOVE (Treasury Vol)", "indice"),
    ("Volatility", "vxeem", "VXEEM Index", "PX_LAST", "VXEEM (EM Vol)", "indice"),
    ("Volatility", "skew", "SKEW Index", "PX_LAST", "SKEW (Tail Risk)", "indice"),
    ("Volatility", "vix_3m", "VIX3M Index", "PX_LAST", "VIX 3M", "indice"),
    ("Volatility", "v2x", "V2X Index", "PX_LAST", "V2X (EuroStoxx Vol)", "indice"),
    ("Volatility", "gvz", "GVZ Index", "PX_LAST", "GVZ (Gold Vol)", "indice"),

    # ── Macro Conditions (6 series) ─────────────────────────────────────
    ("Macro_Conditions", "lei_usa", "LEI TOTL Index", "PX_LAST", "Leading Econ Indicators", "indice"),
    ("Macro_Conditions", "cfnai", "CFNAI Index", "PX_LAST", "Chicago Fed Natl Activity", "indice"),
    ("Macro_Conditions", "nfib_optimism", "SBOITOTL Index", "PX_LAST", "NFIB Small Business", "indice"),
    ("Macro_Conditions", "umich_sentiment", "CONSSENT Index", "PX_LAST", "Michigan Consumer Sent", "indice"),
    ("Macro_Conditions", "ism_new_orders", "NAPMNEWO Index", "PX_LAST", "ISM New Orders", "indice"),
    ("Macro_Conditions", "adp_employment", "ADP CHNG Index", "PX_LAST", "ADP Employment Chg", "k"),

    # ── Chile (5 series) ────────────────────────────────────────────────
    ("Chile", "ipsa", "IPSA Index", "PX_LAST", "IPSA", "indice"),
    ("Chile", "usdclp", "USDCLP Curncy", "PX_LAST", "USD/CLP", "CLP"),
    ("Chile", "bcp_10y", "GICL10YR Index", "PX_LAST", "BCP Chile 10Y", "%"),
    ("Chile", "chile_cpi_yoy", "CLCPIYOY Index", "PX_LAST", "Chile CPI YoY", "%"),
    ("Chile", "tpm_chile", "CHBCRF Index", "PX_LAST", "TPM Chile", "%"),

    # ── Factor Returns (5 series) ─────────────────────────────────────
    ("Factor_Returns", "factor_quality", "M1USQU Index", "PX_LAST", "MSCI USA Quality YTD", "%"),
    ("Factor_Returns", "factor_momentum", "MTUM US Equity", "PX_LAST", "MSCI USA Momentum YTD", "%"),
    ("Factor_Returns", "factor_value", "M1USEV Index", "PX_LAST", "MSCI USA Enh Value YTD", "%"),
    ("Factor_Returns", "factor_growth", "M1US000G Index", "PX_LAST", "MSCI USA Growth YTD", "%"),
    ("Factor_Returns", "factor_size", "M1USSC Index", "PX_LAST", "MSCI USA Small Cap YTD", "%"),

    # ── SOFR Swap Curve (19 series) ────────────────────────────────────
    ("SOFR", "sofr_rate", "SOFRRATE Index", "PX_LAST", "SOFR Overnight Rate", "%"),
    ("SOFR", "sofr_1w", "USOSFR1Z Curncy", "PX_LAST", "SOFR Swap 1W", "%"),
    ("SOFR", "sofr_1m", "USOSFRA Curncy", "PX_LAST", "SOFR Swap 1M", "%"),
    ("SOFR", "sofr_3m", "USOSFRC Curncy", "PX_LAST", "SOFR Swap 3M", "%"),
    ("SOFR", "sofr_6m", "USOSFRF Curncy", "PX_LAST", "SOFR Swap 6M", "%"),
    ("SOFR", "sofr_1y", "USOSFR1 Curncy", "PX_LAST", "SOFR Swap 1Y", "%"),
    ("SOFR", "sofr_2y", "USOSFR2 Curncy", "PX_LAST", "SOFR Swap 2Y", "%"),
    ("SOFR", "sofr_3y", "USOSFR3 Curncy", "PX_LAST", "SOFR Swap 3Y", "%"),
    ("SOFR", "sofr_4y", "USOSFR4 Curncy", "PX_LAST", "SOFR Swap 4Y", "%"),
    ("SOFR", "sofr_5y", "USOSFR5 Curncy", "PX_LAST", "SOFR Swap 5Y", "%"),
    ("SOFR", "sofr_6y", "USOSFR6 Curncy", "PX_LAST", "SOFR Swap 6Y", "%"),
    ("SOFR", "sofr_7y", "USOSFR7 Curncy", "PX_LAST", "SOFR Swap 7Y", "%"),
    ("SOFR", "sofr_8y", "USOSFR8 Curncy", "PX_LAST", "SOFR Swap 8Y", "%"),
    ("SOFR", "sofr_9y", "USOSFR9 Curncy", "PX_LAST", "SOFR Swap 9Y", "%"),
    ("SOFR", "sofr_10y", "USOSFR10 Curncy", "PX_LAST", "SOFR Swap 10Y", "%"),
    ("SOFR", "sofr_15y", "USOSFR15 Curncy", "PX_LAST", "SOFR Swap 15Y", "%"),
    ("SOFR", "sofr_20y", "USOSFR20 Curncy", "PX_LAST", "SOFR Swap 20Y", "%"),
    ("SOFR", "sofr_25y", "USOSFR25 Curncy", "PX_LAST", "SOFR Swap 25Y", "%"),
    ("SOFR", "sofr_30y", "USOSFR30 Curncy", "PX_LAST", "SOFR Swap 30Y", "%"),

    # ── Intl Sovereign Curves (9 series) ──────────────────────────────
    # German Bund
    ("Intl_Curves", "bund_2y", "GTDEM2Y Govt", "PX_LAST", "Bund 2Y Yield", "%"),
    ("Intl_Curves", "bund_5y", "GTDEM5Y Govt", "PX_LAST", "Bund 5Y Yield", "%"),
    ("Intl_Curves", "bund_30y", "GTDEM30Y Govt", "PX_LAST", "Bund 30Y Yield", "%"),
    # UK Gilt
    ("Intl_Curves", "gilt_2y", "GUKG2 Index", "PX_LAST", "Gilt 2Y Yield", "%"),
    ("Intl_Curves", "gilt_5y", "GUKG5 Index", "PX_LAST", "Gilt 5Y Yield", "%"),
    ("Intl_Curves", "gilt_30y", "GUKG30 Index", "PX_LAST", "Gilt 30Y Yield", "%"),
    # JGB
    ("Intl_Curves", "jgb_2y", "GJGB2 Index", "PX_LAST", "JGB 2Y Yield", "%"),
    ("Intl_Curves", "jgb_5y", "GJGB5 Index", "PX_LAST", "JGB 5Y Yield", "%"),
    ("Intl_Curves", "jgb_30y", "GJGB30 Index", "PX_LAST", "JGB 30Y Yield", "%"),
]

# Sheet display order and tab colors
SHEET_CONFIG = [
    ("PMI", "2B6CB0"),
    ("China", "C53030"),
    ("CDS", "DD6B20"),
    ("Credit_Spreads", "DD6B20"),
    ("EM_Spreads", "744210"),
    ("Real_Yields", "DD6B20"),
    ("CPI_Componentes", "2B6CB0"),
    ("EPFR_Flows", "276749"),
    ("Positioning", "276749"),
    ("Valuaciones", "276749"),
    ("Volatility", "8B5CF6"),
    ("Macro_Conditions", "2B6CB0"),
    ("Chile", "0066CC"),
    ("Factor_Returns", "8B5CF6"),
    ("SOFR", "DD6B20"),
    ("Intl_Curves", "DD6B20"),
]


# ═════════════════════════════════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════════════════════════════════

def _month_add(base_date, months_fwd):
    """Add N months to a date, returning (year, month)."""
    y = base_date.year
    m = base_date.month + months_fwd
    while m > 12:
        m -= 12
        y += 1
    return y, m


def _month_subtract(base_date, months_back):
    """Subtract N months from a date, returning (year, month)."""
    y = base_date.year
    m = base_date.month - months_back
    while m <= 0:
        m += 12
        y -= 1
    return y, m


def _auto_width(ws, ncols, min_w=12, max_w=40):
    """Auto-size columns based on content."""
    for c in range(1, ncols + 1):
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(c)].width = max_len


# ═════════════════════════════════════════════════════════════════════════
# TEMPLATE GENERATOR
# ═════════════════════════════════════════════════════════════════════════

def create_template(path: str):
    """Create the Bloomberg Excel template with time series layout."""
    wb = Workbook()
    today = date.today()

    # ── CONFIG sheet ────────────────────────────────────────────────────
    ws_config = wb.active
    ws_config.title = "CONFIG"
    ws_config.sheet_properties.tabColor = "999999"

    ws_config.cell(row=1, column=1, value="Parametro").font = header_font
    ws_config.cell(row=1, column=1).fill = header_fill
    ws_config.cell(row=1, column=2, value="Valor").font = header_font
    ws_config.cell(row=1, column=2).fill = header_fill
    ws_config.cell(row=1, column=3, value="Detalle").font = header_font
    ws_config.cell(row=1, column=3).fill = header_fill

    config_rows = [
        ("ultima_actualizacion", "", "Timestamp del ultimo refresh con Bloomberg"),
        ("series_actualizadas", "", "Cantidad de series en ultimo refresh"),
        ("meses_historia", str(MONTHS_HISTORY), "Meses de historia por serie"),
        ("version_template", "2.0", "Version del template (time series)"),
        ("creado", today.strftime("%Y-%m-%d"), "Fecha de creacion del template"),
    ]
    for i, (param, val, detail) in enumerate(config_rows, 2):
        ws_config.cell(row=i, column=1, value=param).font = normal_font
        ws_config.cell(row=i, column=2, value=val).font = normal_font
        ws_config.cell(row=i, column=3, value=detail).font = note_font

    ws_config.column_dimensions["A"].width = 25
    ws_config.column_dimensions["B"].width = 25
    ws_config.column_dimensions["C"].width = 50

    # ── INSTRUCCIONES sheet ─────────────────────────────────────────────
    ws_inst = wb.create_sheet(title="INSTRUCCIONES")
    ws_inst.sheet_properties.tabColor = ORANGE

    instructions = [
        ("GREYBARK RESEARCH - Bloomberg Data Template v2.0", title_font),
        ("", normal_font),
        ("Este archivo es el puente entre Bloomberg y el sistema de reportes.", subtitle_font),
        ("Contiene 148 series de tiempo con 10 anos de historia mensual.", normal_font),
        ("", normal_font),
        ("COMO ACTUALIZAR (1 click):", subtitle_font),
        ("", normal_font),
        ("1. Abrir este archivo en un PC con Bloomberg Terminal activo", normal_font),
        ("2. Habilitar macros si Excel lo pide", normal_font),
        ("3. Ir a la pestana Desarrollador > Macros > ACTUALIZAR > Ejecutar", normal_font),
        ("4. Esperar ~2-3 minutos (la macro muestra progreso en barra de estado)", normal_font),
        ("5. La macro guarda automaticamente al terminar", normal_font),
        ("", normal_font),
        ("IMPORTANTE:", subtitle_font),
        ("  - Las hojas de datos SOLO contienen valores puros (no formulas)", normal_font),
        ("  - Es SEGURO abrir este archivo en PCs sin Bloomberg", normal_font),
        ("  - La macro crea una hoja temporal _STAGING con formulas BDH,", normal_font),
        ("    copia los valores, y elimina la hoja temporal", normal_font),
        ("  - Frecuencia recomendada: 1x al mes, dia 7-8 del mes", normal_font),
        ("", normal_font),
        ("SETUP INICIAL (solo 1 vez):", subtitle_font),
        ("  1. Abrir el Editor de VBA (Alt+F11)", normal_font),
        ("  2. Archivo > Importar > seleccionar bloomberg_macro.bas", normal_font),
        ("  3. Guardar como .xlsm (Excel con macros)", normal_font),
        ("", normal_font),
        ("HOJAS DE DATOS:", subtitle_font),
        ("", normal_font),
    ]

    for sheet_name, tab_color in SHEET_CONFIG:
        n_series = sum(1 for c in CATALOG if c[0] == sheet_name)
        instructions.append(
            (f"  {sheet_name:<20} {n_series} series", normal_font)
        )

    instructions += [
        ("", normal_font),
        ("LAYOUT DE CADA HOJA:", subtitle_font),
        ("  Fila 1: Descripcion (encabezado visible)", normal_font),
        ("  Fila 2: Campo ID (identificador para el codigo Python)", normal_font),
        ("  Fila 3: Bloomberg Ticker (para la macro VBA)", normal_font),
        ("  Fila 4: BDH Field (PX_LAST, BEST_PE_RATIO, etc.)", normal_font),
        ("  Fila 5+: Datos (columna A = fecha YYYY-MM, columnas B+ = valores)", normal_font),
        ("", normal_font),
        (f"  Generado: {today.strftime('%Y-%m-%d')}", note_font),
    ]

    for i, (text, font) in enumerate(instructions, 1):
        cell = ws_inst.cell(row=i, column=1, value=text)
        cell.font = font
    ws_inst.column_dimensions["A"].width = 80

    # ── DATA SHEETS ─────────────────────────────────────────────────────
    for sheet_name, tab_color in SHEET_CONFIG:
        series = [c for c in CATALOG if c[0] == sheet_name]
        if not series:
            continue

        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = tab_color
        n_series = len(series)

        # Row 1: Descriptions (visible header)
        ws.cell(row=1, column=1, value="").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        for j, (_, _, _, _, desc, unit) in enumerate(series, 2):
            cell = ws.cell(row=1, column=j, value=f"{desc} ({unit})")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 2: Campo IDs (small gray)
        ws.cell(row=2, column=1, value="campo_id").font = small_gray_font
        for j, (_, campo_id, _, _, _, _) in enumerate(series, 2):
            cell = ws.cell(row=2, column=j, value=campo_id)
            cell.font = small_gray_font
            cell.alignment = Alignment(horizontal="center")

        # Row 3: Bloomberg Tickers (small gray)
        ws.cell(row=3, column=1, value="ticker").font = small_gray_font
        for j, (_, _, ticker, _, _, _) in enumerate(series, 2):
            cell = ws.cell(row=3, column=j, value=ticker)
            cell.font = small_gray_font
            cell.alignment = Alignment(horizontal="center")

        # Row 4: BDH Fields (small gray)
        ws.cell(row=4, column=1, value="Fecha").font = subtitle_font
        for j, (_, _, _, field, _, _) in enumerate(series, 2):
            cell = ws.cell(row=4, column=j, value=field)
            cell.font = small_gray_font
            cell.alignment = Alignment(horizontal="center")

        # Row 5+: Date rows (3 months forward + 24 months back, newest first)
        total_rows = MONTHS_FORWARD + MONTHS_HISTORY
        for r in range(total_rows):
            # r=0 → +3 months, r=1 → +2, ..., r=3 → current, r=4 → -1, etc.
            offset = MONTHS_FORWARD - r
            if offset >= 0:
                y, m = _month_add(today, offset)
            else:
                y, m = _month_subtract(today, -offset)
            date_str = f"{y}-{m:02d}"
            row_num = DATA_START_ROW + r

            # Date in column A
            ws.cell(row=row_num, column=1, value=date_str).font = date_font

            # Empty value cells with yellow fill
            for j in range(2, n_series + 2):
                cell = ws.cell(row=row_num, column=j)
                cell.fill = yellow_fill
                cell.number_format = '0.00'
                cell.font = normal_font

            # Alternating row fill
            if r % 2 == 1:
                for j in range(1, n_series + 2):
                    if j > 1:
                        ws.cell(row=row_num, column=j).fill = alt_fill

        # Formatting
        ws.row_dimensions[1].height = 35
        ws.column_dimensions["A"].width = 12
        for j in range(2, n_series + 2):
            ws.column_dimensions[get_column_letter(j)].width = 16
        ws.freeze_panes = "B5"

    # ── CATALOGO reference sheet ────────────────────────────────────────
    ws_cat = wb.create_sheet(title="Catalogo")
    ws_cat.sheet_properties.tabColor = "999999"

    cat_headers = ["Hoja", "Campo ID", "Bloomberg Ticker", "BDH Field",
                   "Descripcion", "Unidad"]
    for c, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, (sheet, campo, ticker, field, desc, unit) in enumerate(CATALOG, 2):
        ws_cat.cell(row=i, column=1, value=sheet).font = normal_font
        ws_cat.cell(row=i, column=2, value=campo).font = normal_font
        ws_cat.cell(row=i, column=3, value=ticker).font = normal_font
        ws_cat.cell(row=i, column=4, value=field).font = normal_font
        ws_cat.cell(row=i, column=5, value=desc).font = normal_font
        ws_cat.cell(row=i, column=6, value=unit).font = normal_font

        if (i - 2) % 2 == 1:
            for c in range(1, 7):
                ws_cat.cell(row=i, column=c).fill = alt_fill

    _auto_width(ws_cat, len(cat_headers))
    ws_cat.freeze_panes = "A2"

    # ── SAVE ────────────────────────────────────────────────────────────
    wb.save(path)
    n_sheets = sum(1 for s in SHEET_CONFIG)
    print(f"  Template creado: {path}")
    print(f"  Hojas de datos: {n_sheets}")
    print(f"  Series totales: {len(CATALOG)}")
    print(f"  Meses historia: {MONTHS_HISTORY}")
    print(f"  Hojas: {', '.join(ws.title for ws in wb.worksheets)}")


# ═════════════════════════════════════════════════════════════════════════
# LIVE TEMPLATE (no macros — BDH formulas directly in cells)
# ═════════════════════════════════════════════════════════════════════════

BLOCK_SIZE = 135  # Rows per series block (120 months + metadata + buffer)

def create_live_template(path: str):
    """Create Bloomberg Excel with BDH formulas directly in cells.
    No macros needed — just open on Bloomberg PC, data loads, save.

    Layout per themed sheet (series stacked vertically):
      Block N (for each series):
        Row base+0: campo_id | description | ticker | field
        Row base+1: =BDH(ticker, field, start, end, "periodicitySelection", "MONTHLY")
                    ↳ BDH spills: col A = dates, col B = values
        [130 rows gap to next block]
    """
    wb = Workbook()
    today = date.today()

    # Date range: 10 years back, 6 months forward buffer
    start_y, start_m = _month_subtract(today, MONTHS_HISTORY)
    start_date = f"{start_m:02d}/01/{start_y}"
    end_y, end_m = _month_add(today, 6)
    end_date = f"{end_m:02d}/28/{end_y}"

    # ── CONFIG sheet ─────────────────────────────────────────────────
    ws_config = wb.active
    ws_config.title = "CONFIG"
    ws_config.sheet_properties.tabColor = "999999"

    ws_config.cell(row=1, column=1, value="Parametro").font = header_font
    ws_config.cell(row=1, column=1).fill = header_fill
    ws_config.cell(row=1, column=2, value="Valor").font = header_font
    ws_config.cell(row=1, column=2).fill = header_fill
    ws_config.cell(row=1, column=3, value="Detalle").font = header_font
    ws_config.cell(row=1, column=3).fill = header_fill

    config_rows = [
        ("ultima_actualizacion", "", "Se actualiza al abrir en Bloomberg"),
        ("version_template", "3.0-live", "Version live (sin macros, BDH directo)"),
        ("meses_historia", str(MONTHS_HISTORY), "Meses de historia"),
        ("formato", "bloques_verticales", "Series apiladas por bloque en cada hoja"),
        ("block_size", str(BLOCK_SIZE), "Filas por bloque de serie"),
        ("creado", today.strftime("%Y-%m-%d"), "Fecha de creacion"),
    ]
    for i, (param, val, detail) in enumerate(config_rows, 2):
        ws_config.cell(row=i, column=1, value=param).font = normal_font
        ws_config.cell(row=i, column=2, value=val).font = normal_font
        ws_config.cell(row=i, column=3, value=detail).font = note_font

    ws_config.column_dimensions["A"].width = 25
    ws_config.column_dimensions["B"].width = 30
    ws_config.column_dimensions["C"].width = 50

    # ── INSTRUCCIONES sheet ──────────────────────────────────────────
    ws_inst = wb.create_sheet(title="INSTRUCCIONES")
    ws_inst.sheet_properties.tabColor = ORANGE

    instructions = [
        ("GREYBARK RESEARCH - Bloomberg Data (Live)", title_font),
        ("", normal_font),
        ("Este archivo tiene formulas BDH nativas.", subtitle_font),
        ("NO necesita macros. Solo abrir, esperar, guardar.", normal_font),
        ("", normal_font),
        ("COMO ACTUALIZAR:", subtitle_font),
        ("", normal_font),
        ("  1. Abrir este archivo en PC con Bloomberg Terminal activo", normal_font),
        ("  2. Las formulas BDH se calculan automaticamente (~2-3 min)", normal_font),
        ("  3. Guardar (Ctrl+S)", normal_font),
        ("  4. Cerrar. Listo.", normal_font),
        ("", normal_font),
        ("IMPORTANTE:", subtitle_font),
        ("  - NO abrir en PC sin Bloomberg (las formulas mostraran #NAME?)", normal_font),
        ("  - Si lo abres sin Bloomberg NO guardes (cierra sin guardar)", normal_font),
        ("  - Frecuencia: 1x al mes, dia 7-8", normal_font),
        ("", normal_font),
        ("LAYOUT:", subtitle_font),
        ("  Cada serie es un bloque vertical de 135 filas:", normal_font),
        ("  Fila 1 del bloque: campo_id | descripcion | ticker | field", normal_font),
        ("  Fila 2 del bloque: =BDH(...) que rellena fechas y valores", normal_font),
        ("", normal_font),
        ("HOJAS:", subtitle_font),
    ]

    for sheet_name, tab_color in SHEET_CONFIG:
        n_series = sum(1 for c in CATALOG if c[0] == sheet_name)
        instructions.append(
            (f"  {sheet_name:<20} {n_series} series", normal_font)
        )

    instructions += [
        ("", normal_font),
        (f"  148 series | 10 anos | Generado: {today.strftime('%Y-%m-%d')}", note_font),
    ]

    for i, (text, font) in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=text).font = font
    ws_inst.column_dimensions["A"].width = 80

    # ── DATA SHEETS with BDH formulas ────────────────────────────────
    for sheet_name, tab_color in SHEET_CONFIG:
        series = [c for c in CATALOG if c[0] == sheet_name]
        if not series:
            continue

        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = tab_color

        for idx, (_, campo_id, ticker, field, desc, unit) in enumerate(series):
            base_row = idx * BLOCK_SIZE + 1

            # Metadata row
            ws.cell(row=base_row, column=1, value=campo_id).font = subtitle_font
            ws.cell(row=base_row, column=2, value=f"{desc} ({unit})").font = normal_font
            ws.cell(row=base_row, column=3, value=ticker).font = small_gray_font
            ws.cell(row=base_row, column=4, value=field).font = small_gray_font

            # BDH formula in next row
            bdh_row = base_row + 1
            formula = (
                f'=BDH("{ticker}","{field}","{start_date}","{end_date}",'
                f'"periodicitySelection","MONTHLY")'
            )
            ws.cell(row=bdh_row, column=1).value = formula

            # Visual hint: color the metadata row
            for c in range(1, 5):
                ws.cell(row=base_row, column=c).fill = header_fill
                ws.cell(row=base_row, column=c).font = Font(
                    name="Segoe UI", bold=True, color=WHITE, size=10
                )

        # Column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 18

    # ── CATALOGO ─────────────────────────────────────────────────────
    ws_cat = wb.create_sheet(title="Catalogo")
    ws_cat.sheet_properties.tabColor = "999999"

    cat_headers = ["Hoja", "Campo ID", "Bloomberg Ticker", "BDH Field",
                   "Descripcion", "Unidad", "Bloque"]
    for c, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    block_counter = {}
    for i, (sheet, campo, ticker, field, desc, unit) in enumerate(CATALOG, 2):
        block_counter[sheet] = block_counter.get(sheet, -1) + 1
        block_num = block_counter[sheet]
        base_row = block_num * BLOCK_SIZE + 1

        ws_cat.cell(row=i, column=1, value=sheet).font = normal_font
        ws_cat.cell(row=i, column=2, value=campo).font = normal_font
        ws_cat.cell(row=i, column=3, value=ticker).font = normal_font
        ws_cat.cell(row=i, column=4, value=field).font = normal_font
        ws_cat.cell(row=i, column=5, value=desc).font = normal_font
        ws_cat.cell(row=i, column=6, value=unit).font = normal_font
        ws_cat.cell(row=i, column=7, value=f"Row {base_row}").font = small_gray_font

    _auto_width(ws_cat, len(cat_headers))
    ws_cat.freeze_panes = "A2"

    # ── SAVE ─────────────────────────────────────────────────────────
    wb.save(path)

    # Fix: openpyxl adds _xll. prefix to unknown functions (BDH).
    # Some Bloomberg versions don't recognize _xll.BDH — strip the prefix.
    _strip_xll_prefix(path)

    print(f"  Live template: {path}")
    print(f"  Hojas: {len(SHEET_CONFIG)} tematicas")
    print(f"  Series: {len(CATALOG)} con BDH directo")
    print(f"  Rango BDH: {start_date} → {end_date}")


def _strip_xll_prefix(xlsx_path: str):
    """Remove _xll. prefix from BDH formulas in the xlsx XML.

    openpyxl auto-prepends _xll. to unknown functions (Bloomberg BDH/BDP/BDS).
    This causes issues in some Bloomberg versions. We patch the raw XML.
    """
    import zipfile
    import shutil
    import tempfile

    temp_path = xlsx_path + '.tmp'

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                # Only patch worksheet XML files
                if item.filename.startswith('xl/worksheets/'):
                    data = data.replace(b'_xll.BDH(', b'BDH(')
                    data = data.replace(b'_xll.BDP(', b'BDP(')
                    data = data.replace(b'_xll.BDS(', b'BDS(')
                zout.writestr(item, data)

    # Replace original
    os.replace(temp_path, xlsx_path)


# ═════════════════════════════════════════════════════════════════════════
# VBA MODULE GENERATOR
# ═════════════════════════════════════════════════════════════════════════

def write_vba_module(path: str):
    """Write the bloomberg_macro.bas VBA module for one-click refresh."""

    # Build sheet names list for VBA
    sheet_names_vba = ', '.join(f'"{s[0]}"' for s in SHEET_CONFIG)

    vba_code = f'''Attribute VB_Name = "Bloomberg_Refresh"
Option Explicit

' =====================================================================
' Greybark Research - Bloomberg One-Click Refresh Macro
' =====================================================================
' This macro:
'   1. Creates a temporary _STAGING sheet
'   2. Writes BDH formulas for all series
'   3. Waits for Bloomberg to resolve
'   4. Copies values to data sheets
'   5. Deletes staging, updates CONFIG, saves
'
' Data sheet layout:
'   Row 3 = Bloomberg Tickers
'   Row 4 = BDH Fields
'   Row 5+ = Data (Col A = dates YYYY-MM)
' =====================================================================

Private Const ROWS_PER_SERIES As Long = 130
Private Const DATA_START_ROW As Long = 5
Private Const TICKER_ROW As Long = 3
Private Const FIELD_ROW As Long = 4
Private Const MAX_WAIT_SECONDS As Long = 300

Sub ACTUALIZAR()
    ' One-click Bloomberg data refresh
    Dim wb As Workbook: Set wb = ThisWorkbook
    Dim dataSheets As Variant
    dataSheets = Array({sheet_names_vba})

    ' Check Bloomberg availability
    If Not IsBloombergAvailable() Then
        MsgBox "Bloomberg Excel Add-in no detectado." & vbCrLf & _
               "Abra Bloomberg Terminal y habilite el Add-in en Excel.", _
               vbExclamation, "Bloomberg Refresh"
        Exit Sub
    End If

    Dim answer As VbMsgBoxResult
    answer = MsgBox("Actualizar datos Bloomberg?" & vbCrLf & _
                    "Esto tomara ~2-3 minutos.", _
                    vbYesNo + vbQuestion, "Bloomberg Refresh")
    If answer <> vbYes Then Exit Sub

    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    Application.StatusBar = "Bloomberg: Preparando..."

    ' Create staging sheet
    Dim wsStaging As Worksheet
    Set wsStaging = GetOrCreateSheet(wb, "_STAGING")
    wsStaging.Cells.Clear

    ' Calculate date range
    Dim startDate As String
    Dim endDate As String
    startDate = Format(DateAdd("m", -{MONTHS_HISTORY}, Date), "mm/dd/yyyy")
    endDate = Format(Date, "mm/dd/yyyy")

    ' Write BDH formulas to staging
    Dim seriesCount As Long: seriesCount = 0
    Dim i As Long

    For i = LBound(dataSheets) To UBound(dataSheets)
        Dim ws As Worksheet
        Set ws = Nothing
        On Error Resume Next
        Set ws = wb.Sheets(dataSheets(i))
        On Error GoTo 0
        If ws Is Nothing Then GoTo NextSheet

        Dim lastCol As Long
        lastCol = ws.Cells(TICKER_ROW, ws.Columns.Count).End(xlToLeft).Column

        Dim col As Long
        For col = 2 To lastCol
            Dim ticker As String
            Dim fld As String
            ticker = Trim(CStr(ws.Cells(TICKER_ROW, col).Value & ""))
            fld = Trim(CStr(ws.Cells(FIELD_ROW, col).Value & ""))

            If Len(ticker) > 0 And Len(fld) > 0 Then
                Dim baseRow As Long
                baseRow = seriesCount * ROWS_PER_SERIES + 1

                ' Metadata in columns D-F (for copy-back)
                wsStaging.Cells(baseRow, 4).Value = dataSheets(i)
                wsStaging.Cells(baseRow, 5).Value = col
                wsStaging.Cells(baseRow, 6).Value = ticker

                ' BDH formula
                wsStaging.Cells(baseRow, 1).Formula = _
                    "=BDH(""" & ticker & """,""" & fld & """,""" & _
                    startDate & """,""" & endDate & _
                    """,""periodicitySelection"",""MONTHLY"")"

                seriesCount = seriesCount + 1
            End If
        Next col
NextSheet:
    Next i

    If seriesCount = 0 Then
        MsgBox "No se encontraron series para actualizar.", vbInformation
        GoTo Cleanup
    End If

    Application.StatusBar = "Bloomberg: Calculando " & seriesCount & " series..."
    Application.Calculation = xlCalculationAutomatic
    Application.CalculateFull

    ' Wait for Bloomberg to resolve all formulas
    Application.StatusBar = "Bloomberg: Esperando respuesta..."
    If Not WaitForBloomberg(wsStaging, seriesCount) Then
        MsgBox "Timeout esperando Bloomberg (" & MAX_WAIT_SECONDS & "s)." & vbCrLf & _
               "Verifique la conexion a Bloomberg Terminal.", _
               vbExclamation, "Bloomberg Refresh"
        GoTo Cleanup
    End If

    ' Copy values from staging to data sheets
    Application.StatusBar = "Bloomberg: Copiando valores..."
    Dim copied As Long: copied = 0
    Dim s As Long

    For s = 0 To seriesCount - 1
        Dim bRow As Long: bRow = s * ROWS_PER_SERIES + 1
        Dim sheetName As String: sheetName = CStr(wsStaging.Cells(bRow, 4).Value)
        Dim colIdx As Long: colIdx = CLng(wsStaging.Cells(bRow, 5).Value)

        Dim wsTarget As Worksheet
        Set wsTarget = wb.Sheets(sheetName)

        ' Read BDH results: col A = dates, col B = values
        Dim r As Long
        For r = 0 To ROWS_PER_SERIES - 2
            Dim cellDate As Variant
            Dim cellVal As Variant
            cellDate = wsStaging.Cells(bRow + r, 1).Value
            cellVal = wsStaging.Cells(bRow + r, 2).Value

            If IsEmpty(cellDate) Then Exit For
            If Not IsDate(cellDate) Then Exit For

            ' Find matching row in target sheet (by year-month)
            Dim targetRow As Long
            targetRow = FindDateRow(wsTarget, CDate(cellDate))

            If targetRow > 0 And Not IsEmpty(cellVal) Then
                If IsNumeric(cellVal) Then
                    wsTarget.Cells(targetRow, colIdx).Value = CDbl(cellVal)
                    copied = copied + 1
                End If
            End If
        Next r
    Next s

Cleanup:
    ' Delete staging sheet
    Application.DisplayAlerts = False
    On Error Resume Next
    wb.Sheets("_STAGING").Delete
    On Error GoTo 0
    Application.DisplayAlerts = True

    ' Update CONFIG
    On Error Resume Next
    wb.Sheets("CONFIG").Cells(2, 2).Value = Format(Now, "yyyy-mm-dd hh:mm:ss")
    wb.Sheets("CONFIG").Cells(3, 2).Value = seriesCount
    On Error GoTo 0

    ' Save
    wb.Save

    Application.StatusBar = False
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic

    MsgBox "Actualizacion completada" & vbCrLf & _
           "Series: " & seriesCount & vbCrLf & _
           "Valores copiados: " & copied & vbCrLf & _
           Format(Now, "dd/mm/yyyy hh:mm"), _
           vbInformation, "Bloomberg Refresh"
End Sub


Private Function FindDateRow(ws As Worksheet, targetDate As Date) As Long
    ' Find the row matching a given month (YYYY-MM format in col A)
    Dim r As Long
    For r = DATA_START_ROW To DATA_START_ROW + 130
        Dim cellVal As Variant
        cellVal = ws.Cells(r, 1).Value
        If IsEmpty(cellVal) Then Exit For

        ' Try as date object
        If IsDate(cellVal) Then
            If Year(CDate(cellVal)) = Year(targetDate) And _
               Month(CDate(cellVal)) = Month(targetDate) Then
                FindDateRow = r
                Exit Function
            End If
        End If

        ' Try as YYYY-MM string
        Dim s As String: s = CStr(cellVal)
        If Len(s) >= 7 And InStr(s, "-") > 0 Then
            Dim parts() As String
            parts = Split(s, "-")
            If UBound(parts) >= 1 Then
                On Error Resume Next
                Dim y As Long: y = CLng(parts(0))
                Dim m As Long: m = CLng(parts(1))
                On Error GoTo 0
                If y = Year(targetDate) And m = Month(targetDate) Then
                    FindDateRow = r
                    Exit Function
                End If
            End If
        End If
    Next r
    FindDateRow = 0
End Function


Private Function WaitForBloomberg(wsStaging As Worksheet, seriesCount As Long) As Boolean
    Dim elapsed As Long: elapsed = 0

    Do While elapsed < MAX_WAIT_SECONDS
        DoEvents
        Application.Wait Now + TimeSerial(0, 0, 2)
        elapsed = elapsed + 2

        ' Check if any cell still shows Bloomberg requesting status
        Dim stillWaiting As Boolean: stillWaiting = False
        Dim s As Long
        For s = 0 To seriesCount - 1
            Dim checkRow As Long: checkRow = s * ROWS_PER_SERIES + 1
            Dim cellText As String
            On Error Resume Next
            cellText = UCase(CStr(wsStaging.Cells(checkRow, 1).Text))
            On Error GoTo 0

            If InStr(cellText, "REQUESTING") > 0 Or _
               InStr(cellText, "#GETTING") > 0 Then
                stillWaiting = True
                Exit For
            End If
        Next s

        If Not stillWaiting Then
            WaitForBloomberg = True
            Exit Function
        End If

        Application.StatusBar = "Bloomberg: Esperando... " & elapsed & "s"
    Loop

    WaitForBloomberg = False
End Function


Private Function GetOrCreateSheet(wb As Workbook, sheetName As String) As Worksheet
    On Error Resume Next
    Set GetOrCreateSheet = wb.Sheets(sheetName)
    On Error GoTo 0

    If GetOrCreateSheet Is Nothing Then
        Set GetOrCreateSheet = wb.Sheets.Add(After:=wb.Sheets(wb.Sheets.Count))
        GetOrCreateSheet.Name = sheetName
    End If
    GetOrCreateSheet.Visible = xlSheetVisible
End Function


Private Function IsBloombergAvailable() As Boolean
    Dim addIn As COMAddIn
    On Error Resume Next
    For Each addIn In Application.COMAddIns
        If InStr(1, addIn.progID, "Bloomberg", vbTextCompare) > 0 Then
            If addIn.Connect Then
                IsBloombergAvailable = True
                Exit Function
            End If
        End If
    Next
    On Error GoTo 0
    IsBloombergAvailable = False
End Function
'''

    with open(path, 'w', encoding='utf-8') as f:
        f.write(vba_code)

    print(f"  VBA module: {path}")


# ═════════════════════════════════════════════════════════════════════════
# SETUP SCRIPT GENERATOR (one-time: imports VBA + adds button)
# ═════════════════════════════════════════════════════════════════════════

def write_setup_script(path: str, xlsx_name: str, bas_name: str):
    """Generate setup_bloomberg.vbs — one-time script that creates the .xlsm
    with embedded VBA and a big ACTUALIZAR button on the INSTRUCCIONES sheet."""

    vbs_code = r'''
' =====================================================================
' Greybark Research - Bloomberg Setup (correr UNA sola vez)
' =====================================================================
' Este script:
'   1. Abre bloomberg_data.xlsx
'   2. Importa la macro VBA
'   3. Agrega un boton ACTUALIZAR en la hoja INSTRUCCIONES
'   4. Guarda como .xlsm (con macros)
'   5. Cierra todo
'
' Despues de esto, tu amigo solo abre el .xlsm y hace click en el boton.
' =====================================================================

Option Explicit

Dim fso, scriptDir, xlsxPath, basPath, xlsmPath
Dim xlApp, wb, wsInst, btn, shp

Set fso = CreateObject("Scripting.FileSystemObject")
scriptDir = fso.GetParentFolderName(WScript.ScriptFullName)

xlsxPath = scriptDir & "\" & "XLSX_NAME"
basPath  = scriptDir & "\" & "BAS_NAME"
xlsmPath = scriptDir & "\" & "XLSM_NAME"

' Check files exist
If Not fso.FileExists(xlsxPath) Then
    MsgBox "No se encontro: " & xlsxPath, vbCritical, "Setup Bloomberg"
    WScript.Quit
End If
If Not fso.FileExists(basPath) Then
    MsgBox "No se encontro: " & basPath, vbCritical, "Setup Bloomberg"
    WScript.Quit
End If

' Delete old .xlsm if exists
If fso.FileExists(xlsmPath) Then
    fso.DeleteFile xlsmPath, True
End If

MsgBox "Setup Bloomberg" & vbCrLf & vbCrLf & _
       "Esto va a:" & vbCrLf & _
       "  1. Abrir el Excel" & vbCrLf & _
       "  2. Importar la macro" & vbCrLf & _
       "  3. Agregar boton ACTUALIZAR" & vbCrLf & _
       "  4. Guardar como .xlsm" & vbCrLf & vbCrLf & _
       "Toma ~10 segundos. Click OK para continuar.", _
       vbInformation, "Greybark Research - Setup"

' Open Excel
Set xlApp = CreateObject("Excel.Application")
xlApp.Visible = False
xlApp.DisplayAlerts = False

' Need to enable VBA access (Trust Center setting)
On Error Resume Next
xlApp.AutomationSecurity = 1  ' msoAutomationSecurityLow
On Error GoTo 0

Set wb = xlApp.Workbooks.Open(xlsxPath)

' Import VBA module
On Error Resume Next
wb.VBProject.VBComponents.Import basPath
If Err.Number <> 0 Then
    Err.Clear
    ' Try alternative: if Trust Access to VBA is disabled
    MsgBox "No se pudo importar la macro automaticamente." & vbCrLf & vbCrLf & _
           "Debes habilitar acceso a VBA:" & vbCrLf & _
           "  Excel > Archivo > Opciones > Centro de Confianza > " & vbCrLf & _
           "  Config del Centro de Confianza > Configuracion de Macros > " & vbCrLf & _
           "  Marcar 'Confiar en el acceso al modelo de objetos de proyectos de VBA'" & vbCrLf & vbCrLf & _
           "Luego vuelve a ejecutar este script.", _
           vbExclamation, "Setup Bloomberg"
    wb.Close False
    xlApp.Quit
    Set xlApp = Nothing
    WScript.Quit
End If
On Error GoTo 0

' Go to INSTRUCCIONES sheet
Set wsInst = wb.Sheets("INSTRUCCIONES")

' Add a big button (Form Control)
' Parameters: Left, Top, Width, Height
Set btn = wsInst.Buttons.Add(50, 30, 350, 80)
With btn
    .Caption = "ACTUALIZAR BLOOMBERG"
    .OnAction = "ACTUALIZAR"
    .Font.Size = 18
    .Font.Bold = True
    .Font.Name = "Segoe UI"
End With

' Also add a smaller note below the button
wsInst.Shapes.AddTextbox(1, 50, 120, 350, 30).TextFrame.Characters.Text = _
    "Click el boton para actualizar. Toma ~2-3 minutos."
wsInst.Shapes(wsInst.Shapes.Count).TextFrame.Characters.Font.Size = 10
wsInst.Shapes(wsInst.Shapes.Count).TextFrame.Characters.Font.Name = "Segoe UI"
wsInst.Shapes(wsInst.Shapes.Count).TextFrame.Characters.Font.Color = RGB(100, 100, 100)
wsInst.Shapes(wsInst.Shapes.Count).Fill.Visible = False
wsInst.Shapes(wsInst.Shapes.Count).Line.Visible = False

' Move the existing instructions text down
' (shift existing content in column A down by inserting rows at top)
wsInst.Rows("1:9").Insert

' Re-label row 1 with big title
wsInst.Cells(1, 1).Value = ""
wsInst.Cells(1, 1).Font.Size = 14

' Save as .xlsm (52 = xlOpenXMLWorkbookMacroEnabled)
wb.SaveAs xlsmPath, 52
wb.Close False

xlApp.Quit
Set xlApp = Nothing

MsgBox "Setup completado!" & vbCrLf & vbCrLf & _
       "Archivo creado: " & xlsmPath & vbCrLf & vbCrLf & _
       "Tu amigo solo necesita:" & vbCrLf & _
       "  1. Abrir " & fso.GetFileName(xlsmPath) & vbCrLf & _
       "  2. Habilitar macros si Excel lo pide" & vbCrLf & _
       "  3. Click en el boton ACTUALIZAR BLOOMBERG" & vbCrLf & vbCrLf & _
       "Listo! Este script no se necesita de nuevo.", _
       vbInformation, "Greybark Research - Setup"
'''

    xlsm_name = xlsx_name.replace('.xlsx', '.xlsm')
    vbs_code = vbs_code.replace('XLSX_NAME', xlsx_name)
    vbs_code = vbs_code.replace('BAS_NAME', bas_name)
    vbs_code = vbs_code.replace('XLSM_NAME', xlsm_name)

    with open(path, 'w', encoding='utf-8') as f:
        f.write(vbs_code)

    print(f"  Setup script: {path}")


# ═════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    base_dir = Path(__file__).parent
    input_dir = base_dir / "input"
    input_dir.mkdir(exist_ok=True)

    print("\nGreybark Research - Bloomberg Template Generator")
    print("=" * 55)

    # Version LIVE (sin macros — BDH directo) ← la principal
    live_path = str(input_dir / "bloomberg_live.xlsx")
    create_live_template(live_path)

    # Version con macro (backup)
    xlsx_path = str(input_dir / "bloomberg_data.xlsx")
    bas_path = str(input_dir / "bloomberg_macro.bas")
    create_template(xlsx_path)
    write_vba_module(bas_path)

    print(f"\n  Para tu amigo: bloomberg_live.xlsx")
    print(f"  Solo abrir en Bloomberg → esperar → Ctrl+S → cerrar")
    print(f"  (backup con macro: bloomberg_data.xlsx + bloomberg_macro.bas)")
